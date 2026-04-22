import winsound
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import openpyxl
import datetime
import serial
import time
import openpyxl
from instrument import uxr0134a
from instrument import DM3058E

def test():
   wb=openpyxl.load_workbook("qpll_kvco.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   address = "TCPIP0::169.254.202.222::inst0::INSTR"
   dm=DM3058E.Dm3058e(address="USB0::0x1AB1::0x09C4::DM3L273100483::INSTR")
   uxr=uxr0134a.Uxr0134a(address)
   channal=3
   chip_num="SS01"
   temp_volt="NTNV"
   lane=0
   txt_file="P5_TEST_TXT/ate_nearpma_PCIE_32GBPS_QPLL0_direct_r260306.txt"
   configure_from_txt(txt_file,device1)
   # 断开Qpll环路
   write_serdes_lane_register(device1, lane, 0xf09f, 0xbd8b)
   # 设置ATESTP到vctrl
   write_serdes_lane_register(device1, lane, 0xf05c, 0x0400)
   # 设置固定的vctrl
   write_serdes_lane_register(device1, lane, 0xf05b, 0x0200)
   write_serdes_bit_reg(device1, lane, 0xf05b, 0, 0, 0)
   write_serdes_bit_reg(device1, lane, 0xf05b, 1, 1, 0)
   Vctrl_mid=dm.get_voltage()
   uxr.set_uxr_state("RUN")
   time.sleep(0.1)
   uxr.auto_verticalscale(channal)
   time.sleep(0.1)
   uxr.clear_display()
   time.sleep(0.1)
   uxr.set_uxr_state("stop")
   freq_mid = uxr.get_meas_val(channal, "FREQuency", "MEAN")
   freq_mid = round(float(freq_mid) * 1e-9, 3)
   print(f"此时频率是{freq_mid}Ghz,此时Vctrl_mid的电压是{Vctrl_mid}V")
   # high
   write_serdes_bit_reg(device1, lane, 0xf05b, 0, 0, 1)
   # low
   write_serdes_bit_reg(device1, lane, 0xf05b, 1, 1, 0)
   Vctrl_hi = dm.get_voltage()
   uxr.set_uxr_state("RUN")
   time.sleep(0.1)
   uxr.auto_verticalscale(channal)
   time.sleep(0.1)
   uxr.clear_display()
   time.sleep(0.1)
   uxr.set_uxr_state("stop")
   freq_hi = uxr.get_meas_val(channal, "FREQuency", "MEAN")
   freq_hi = round(float(freq_hi) * 1e-9, 3)
   print(f"此时频率是{freq_hi}Ghz,此时Vctrl_hi的电压是{Vctrl_hi}V")
   # high
   write_serdes_bit_reg(device1, lane, 0xf05b, 0, 0, 0)
   # low
   write_serdes_bit_reg(device1, lane, 0xf05b, 1, 1, 1)
   Vctrl_lo = dm.get_voltage()
   uxr.set_uxr_state("RUN")
   time.sleep(0.1)
   uxr.auto_verticalscale(channal)
   time.sleep(0.1)
   uxr.clear_display()
   time.sleep(0.1)
   uxr.set_uxr_state("stop")
   freq_lo = uxr.get_meas_val(channal, "FREQuency", "MEAN")
   freq_lo = round(float(freq_lo) * 1e-9, 3)
   print(f"此时频率是{freq_lo}Ghz,此时Vctrl_lo的电压是{Vctrl_lo}V")
   temp_list = [chip_num, temp_volt,freq_mid,Vctrl_mid,freq_hi,Vctrl_hi, freq_lo,Vctrl_lo]
   for i in range(len(temp_list)):
       ws.cell(row=last_row, column=i + 1, value=temp_list[i])
       wb.save("qpll_kvco.xlsx")






def soc_init( port='COM21', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        flag=0
        top_sequences=["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                       "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"]
        for i, seq in enumerate(top_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                    big_endian_value = int.from_bytes(response[:4], byteorder='big')  # 大端解析
                    little_endian_bytes = big_endian_value.to_bytes(4, byteorder='little')  # 转为小端字节
                    print(f"接收数据: {little_endian_bytes.hex()}")
                else:
                    print("[警告] 未收到响应数据")
                    print("soc init failed")
                    flag = 1
                    break

        if(flag==0):
            print("===========soc init success==========")
    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def phy_init(device):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0xf01e)
    print("0xf01e:", value)
    if (value == "0x4967"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()


def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")



if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    port=get_com_port_for_channel("C")
    soc_init(port=port, baudrate=115200)
    FTDI_CABLE_SERDES = b"A"  #
    device1 = setup_device(FTDI_CABLE_SERDES)
    phy_init(device1)
    test()
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    close_serdes(device1)
    print(f"总耗时: {duration}")
    play_sound_test_complete()