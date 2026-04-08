import openpyxl
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import datetime






def save_register_data_to_excel(txt_file, device, lane, external_times, excel_file="DC.xlsx"):
    """
    读取txt文件中的寄存器信息，测量一次并保存到Excel，追加一行。
    external_times: 外部循环次数，用于填充times列。
    """
    # 解析寄存器信息（复用之前的 parse_register_txt 函数）
    register_info = parse_register_txt(txt_file, lane)
    if not register_info:
        print("没有找到任何寄存器信息")
        return

    # 检查文件是否存在，决定是否写入表头
    file_exists = os.path.isfile(excel_file)
    if file_exists:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.worksheets[0]
        last_row = ws.max_row+1
        if external_times==0:
            # 写入表头（寄存器名称）
            ws.cell(row=last_row, column=1, value="times")
            for col, info in enumerate(register_info, start=2):
                ws.cell(row=last_row, column=col, value=info['comment'])
            last_row += 1

            # 写入地址行
            ws.cell(row=last_row, column=1, value="")
            for col, info in enumerate(register_info, start=2):
                addr_str = f"0x{info['actual_addr']:04X}"
                ws.cell(row=last_row, column=col, value=addr_str)
            last_row += 1

            # 写入位段行
            ws.cell(row=last_row, column=1, value="")
            for col, info in enumerate(register_info, start=2):
                bits_str = f"[{info['high_bit']}:{info['low_bit']}]"
                ws.cell(row=last_row, column=col, value=bits_str)
            last_row += 1


        # 读取当前值并写入一行数据
        values = []
        for info in register_info:
            raw = read_serdes_register(device, lane, info['actual_addr'])
            raw = int(raw, 16)
            width = info['high_bit'] - info['low_bit'] + 1
            mask = (1 << width) - 1
            bit_val = (raw >> info['low_bit']) & mask
            values.append(bit_val)

        ws.cell(row=last_row, column=1, value=external_times)
        for col, val in enumerate(values, start=2):
            ws.cell(row=last_row, column=col, value=f"0x{val:04X}")
        wb.save(excel_file)
        print(f"数据已追加到 {excel_file}，外部循环次数: {external_times}")


# 使用示例（需先定义 read_serdes_register 函数）
if __name__ == "__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    full_lane = False
    lane = 0  # 逻辑通道号
    soc_init(port='COM25', baudrate=115200)
    for times in range(5):
        phy_reset(port='COM25', baudrate=115200)
        lane_reset(port='COM25', baudrate=115200)
        release_phy_reset(port='COM25', baudrate=115200)
        FTDI_CABLE_SERDES = b"A"  #
        device1 = setup_device(FTDI_CABLE_SERDES)
        phy_init(device1)
        time.sleep(0.1)
        err_count = read_serdes_register(device1, 0, 0xe057)
        txt_file = 'PCIE_32_QPLL0_direct_gui_contin_off.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        txt_file = 'cont_os_cal_final.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        txt_file = 'AFE_DFE_adapt_mode4.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        write_serdes_lane_register(device1, lane, 0xe058, 0x1c00)
        time.sleep(1)
        write_serdes_lane_register(device1, lane, 0xe058, 0x1c00)
        time.sleep(1)
        release_lane_reset(port='COM25', baudrate=115200)
        time.sleep(0.1)
        if full_lane == True:
            for lane in range(4):
                txt_file = "phy_init_cal_eq_code_rd_final.txt"
                save_register_data_to_excel(txt_file, device1, lane, external_times=times, excel_file="read_eq_code.xlsx")
        else:
            txt_file = "phy_init_cal_eq_code_rd_final.txt"
            save_register_data_to_excel(txt_file, device1, lane, external_times=times, excel_file="read_eq_code.xlsx")
        close_serdes(device1)



