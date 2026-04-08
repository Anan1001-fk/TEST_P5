import winsound
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import openpyxl
from instrument import DM3058E
import datetime
import serial
import tkinter as tk
from tkinter import messagebox
from plyer import notification


test_programs_config_p = {
   "rx_vp": {
        "registers": [ (0xe0f1,0x4000), (0xe0f1,0xc000)],
    },
    "rx_vref_vco": {
        "registers": [ (0xe0f1,0x4000), (0xe0eb,0x8000)],
    },
    "rx_vref_osc_ref": {
        "registers": [ (0xe0eb,0x4000)],
    },
    "rx_vreg_dfe_byp": {
        "registers": [ (0xe0eb,0x2000)],
    },
    "rx_vreg_dfe": {
        "registers": [ (0xe0eb,0x1000)],
    },
    "rx_vreg_clk": {
        "registers": [ (0xe0eb,0x0800)],
    },
    "rx_vdd": {
        "registers": [ (0xe0eb,0x0400)],
    },
    "rx_gd": {
        "registers": [ (0xe0eb,0x0200)],
    },
    "rx_vbp": {
        "registers": [(0xe0eb,0x0000), (0xe0f0,0x1300)],
    },
    "rx_vref_los": {
        "registers": [(0xe0f0,0x0300), (0xe0ec,0x0800)],
    },
    "rx_vref_fil": {
        "registers": [ (0xe0ec,0x1000)],
    },
    "rx_bp1": {
        "registers": [ (0xe0ec,0x2000)],
    },
    "rx_bp1c": {
        "registers": [ (0xe0ec,0x4000)],
    },
    "rx_vregbg": {
        "registers": [(0xe0ec,0x0000), (0xe0e6,0x0100)],
    },
    "rx_vref_ext": {
        "registers": [ (0xe0e6,0x0200)],
    },
    "rx_ref1": {
        "registers": [ (0xe0e6,0x0400)],
    },
    "rx_ipref": {
        "registers": [ (0xe0e6,0x1000)],
    },
    "rx_inref": {
        "registers": [ (0xe0e6,0x2000)],
    },
    "rx_dfe_even_sum_out": {
        "registers": [(0xe0e6,0x0000), (0xe0f0,0x0700)],
    },
    "rx_dfe_odd_sum_out": {
        "registers": [(0xe0f0,0x0b00)],
    },
    "rx_comp_mux_va1/vc1": {
        "registers": [(0xe0f0,0x0300), (0xe0f1,0x4100)],
    },
    "rx_comp_mux_vb1/vd1": {
        "registers": [ (0xe0f1,0x4200)],
    },
    "tx_pbias": {
        "registers": [ (0xe0f1,0x0000),(0xe0fa,0x4000),(0xe0fb,0x0100)],
    },
    "tx_nbias": {
        "registers": [ (0xe0fb,0x0200)],
    },
    "tx_rxdetref": {
        "registers": [(0xe0fb, 0x0400)],
    },
    "tx_vcm": {
        "registers": [ (0xe0fb,0x8000)],
    },
    "tx_txsp": {
        "registers": [ (0xe0fb,0x2000)],
    },
    "tx_bias_vptx": {
        "registers": [ (0xe0fb,0x0000),(0xe0f5,0x0200)],
    },
    "tx_vdd": {
        "registers": [ (0xe0f5,0x0000),(0xe0fc,0x0100)],
    },
    "tx_gd": {
        "registers": [(0xe0fc, 0x4000)],
    },
    "tx_dcc_cal_mux": {
        "registers": [(0xe0fc, 0x2000)],
    },
    "tx_dcc_comp": {
        "registers": [(0xe0fc, 0x1000)],
    },
    "tx_vptx": {
        "registers": [(0xe0fc, 0x0800)],
    },
    "tx_vreg_tx": {
        "registers": [(0xe0fc, 0x0000),(0xe0f8,0x0100)],
    },
    "tx_vreg_vptx": {
        "registers": [(0xe0f8, 0x1000)],
    },
    "tx_dcc_vdac_diff": {
        "registers": [(0xe0f8, 0x0000),(0xe0ff, 0x2000)],
    },
    "tx_dcc_vdac_com": {
        "registers": [(0xe0ff, 0x1000)],
    },
    "tx_vboost": {
        "registers": [(0xe0ff, 0x0000), (0xe0fa, 0x4800)],
    },
    "tx_vboost_vref": {
        "registers": [(0xe0fa, 0x5400)],
    },
    "tx_vph_half": {
        "registers": [(0xe0fa, 0xc000)],
    },
    "txrx_close": {
        "registers": [(0xe0f1, 0x0000), (0xe0fa, 0x0000)],
    },
    "cpll_vreg_refclk": {
        "registers": [(0xe0fa, 0x8b00), (0xe0f0, 0x0b00), (0xe0fa, 0xab00)],
    },
    "cpll_vreg_outclk": {
        "registers": [(0xe0fa, 0x9b00)],
    },
    "cpll_clkpi": {
        "registers": [(0xe0fa, 0x8b00),(0xe0f0, 0x8b00)],
    },
    "cpll_clkmux": {
        "registers": [(0xe0f0, 0x4b00)],
    },
    "cpll_vcobuf": {
        "registers": [(0xe0f0, 0x2b00)],
    },
    "cpll_cp": {
        "registers": [(0xe0f0, 0x0b00),(0xe0f5, 0x0400)],
    },
    "cpll_v2i": {
        "registers": [(0xe0f5, 0x0100)],
    },
    "cpll_close": {
        "registers": [(0xe0f5,0x0000),(0xe0fa,0x0b00)],
    },
    "qpll_dll_avcc": {
        "registers": [(0xf05b, 0x0200), (0xf050, 0x0800)],
    },
    "qpll_vddh": {
        "registers": [(0xf050, 0x0200)],
    },
    "qpll_pfd_avcc": {
        "registers": [(0xf050, 0x0100)],
    },
    "qpll_vreg_right": {
        "registers": [(0xf050, 0x0000), (0xf05c, 0x8000)],
    },
    "qpll_vreg_outclk": {
        "registers": [(0xf05c, 0x4000)],
    },
    "qpll_vreg_cp": {
        "registers": [(0xf05c, 0x2000)],
    },
    "qpll_avcc": {
        "registers": [(0xf05c, 0x1000)],
    },
    "qpll_vreg_vco": {
        "registers": [(0xf05c, 0x0000), (0xf05d, 0x0400)],
    },
    "qpll_vreg_left": {
        "registers": [(0xf05d, 0x0100)],
    },
    "qpll_close": {
        "registers": [(0xf05d, 0x0000), (0xf05b, 0x0000)],
    },
    "common_inbuf_vref_vp": {
        "registers": [(0xf051, 0x2000), (0xf051, 0xa100)],
    },
    "common_vref_repeat_clk": {
        "registers": [(0xf051, 0x2100), (0xf04a, 0x6000)],
    },
    "common_bias_pll_vref": {
        "registers": [(0xf051, 0x0000), (0xf04a, 0x4000),(0xf054, 0x8000)],
    },
    "common_bias_avcch": {
        "registers": [(0xf054, 0x4000)]},
    "common_bias_bg_vref_rx": {
        "registers": [(0xf054, 0x2000)]},
    "common_bias_avcc_rx": {
        "registers": [(0xf054, 0x1000)]},
    "common_bias_avcc/2": {
        "registers": [(0xf054, 0x400)]},
    "common_bias_vbg": {
        "registers": [(0xf054, 0x100)]},
    "common_bias_bg_vref_tx": {
        "registers": [(0xf054, 0x0000),(0xf052, 0x2000)]},
    "close_common": {
        "registers": [(0xf052, 0x0000)]},

}

test_programs_config_n = {
    "rx vbn": {
        "registers": [(0xe0f1, 0x4000), (0xe0f0,0x2300)],
    },
    "rx_vco_vosc": {
        "registers": [(0xe0f0, 0x0300), (0xe0e6,0x0800)],
    },
    "rx_gd": {
        "registers": [(0xe0e6, 0x4000)],
    },
    "tx_txsm": {
        "registers": [(0xe0e6, 0x0000), (0xe0fa,0x4000),(0xe0fb,0x4000)],},
    "tx_close": {
        "registers": [(0xe0fa,0x0000),(0xe0fb,0x0000)],
    },
     "qpll_avss": {
        "registers": [(0xf056, 0x0200), (0xf04f, 0x0400)],
    },
    "common_bias_avss": {
        "registers": [(0xf056, 0x0000), (0xf04f, 0x0000),(0xf054, 0x0800)],},
    "common_bias_vbe": {
        "registers": [(0xf054, 0x0100)],},
    "close_common": {
        "registers": [(0xf054, 0x0000)]},

}

def test(device):
   wb=openpyxl.load_workbook("P5_DC.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   chip_num = "TT01"
   temp_volt = "NTNV"
   # 检查硬件先测P端还是N端
   polarity = "p"  # "N"  "P"
   tk.messagebox.showinfo(title="重要提醒", message=f"请将万用表接在{polarity}端")
   print(f"目前测试: {polarity}")
   value_list = []
   dm=DM3058E.Dm3058e(address="USB0::0x1AB1::0x09C4::DM3R261200875::INSTR")
   dm.reset()
   time.sleep(1)
   if(polarity.upper()=="P"):
       temp_list = [chip_num, temp_volt, polarity]
       for i in range(0, 7, 2):
           flag=1
           lane_p=i
           lane=lane_p/2
           for test_program, config in test_programs_config_p.items():
               if (test_program.startswith("cpll")):
                   if(flag==1):
                      lane_p = lane_p + 1
                      flag=0
               if (test_program.startswith("qpll")|test_program.startswith("common")):
                   lane_p=0
               if((test_program=="txrx_close")|(test_program == "cpll_close")|(test_program == "qpll_close")):
                   for reg_addr, reg_value in config["registers"]:
                       write_serdes_lane_register(device1, lane_p, reg_addr, reg_value)
                   continue
               if (test_program == "close_common"):
                   for reg_addr, reg_value in config["registers"]:
                       write_serdes_lane_register(device1, lane_p, reg_addr, reg_value)
                   break
               value_list.append(int(lane))
               value_list.append(test_program)
               print("目前测试lane", int(lane))
               print("目前测试", test_program)
               for reg_addr, reg_value in config["registers"]:
                   write_serdes_lane_register(device1, lane_p, reg_addr, reg_value)
               volt = dm.get_voltage()
               time.sleep(0.1)
               print("电压是", volt, "v")
               value_list.append(volt)
               time.sleep(0.1)
               for i in range(len(temp_list)):
                   ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               for i in range(len(value_list)):
                   ws.cell(row=last_row, column=i + 4, value=value_list[i])
               wb.save("P5_DC.xlsx")
               save_test_data_to_txt("DC", temp_list, value_list)
               last_row=last_row+1
               value_list.clear()
   else:
       temp_list = [chip_num, temp_volt, polarity]
       for i in range(0, 7, 2):
           lane_n = i
           lane = lane_n / 2
           for test_program, config in test_programs_config_n.items():
               if (test_program.startswith("qpll") | test_program.startswith("common")):
                   lane_n = 0
               if (test_program == "tx_close"):
                   for reg_addr, reg_value in config["registers"]:
                       write_serdes_lane_register(device1, lane_n, reg_addr, reg_value)
                   continue
               if (test_program == "close_common"):
                   for reg_addr, reg_value in config["registers"]:
                       write_serdes_lane_register(device1, lane_n, reg_addr, reg_value)
                   break
               value_list.append(int(lane))
               value_list.append(test_program)
               print("目前测试lane", int(lane))
               print("目前测试", test_program)
               for reg_addr, reg_value in config["registers"]:
                   write_serdes_lane_register(device1, lane_n, reg_addr, reg_value)
               volt = dm.get_voltage()
               time.sleep(0.1)
               print("电压是", volt, "v")
               value_list.append(volt)
               time.sleep(0.1)
               for i in range(len(temp_list)):
                   ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               for i in range(len(value_list)):
                   ws.cell(row=last_row, column=i + 4, value=value_list[i])
               wb.save("P5_DC.xlsx")
               save_test_data_to_txt("DC", temp_list, value_list)
               last_row = last_row + 1
               value_list.clear()






def save_test_data_to_txt(filename, temp_list, value_list, timestamp=None):
    """
    保存测试数据到TXT文件

    参数:
        filename: 文件名
        temp_list: [芯片编号, 温度/电压, 极性]
        value_list: [测试通道, 测试程序, 电压值]
        timestamp: 时间戳（可选）
    """
    import datetime

    # 如果没有提供时间戳，生成当前时间
    if timestamp is None:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 合并所有数据
    all_data = temp_list + value_list + [timestamp]

    # 转换为字符串
    data_line = ','.join(str(item) for item in all_data)

    # 检查文件是否存在，决定是否添加表头
    if not os.path.exists(filename):
        # 文件不存在，创建并添加表头
        header = "芯片编号,温度/电压,极性,测试通道,测试程序,电压值,时间戳"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(header + '\n')
            f.write(data_line + '\n')
    else:
        # 文件存在，追加数据
        with open(filename, 'a', encoding='utf-8') as f:
            f.write(data_line + '\n')

    print(f"数据已保存到: {filename}")
    return True

def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")

def conditional_serial_communication(flag,code,port='COM5', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        # 这个是为了让cfg top&phy的寄存器只走一次
        if flag==1:
            # cfg top通路
            top_sequences = [
            '1104b040485000000000',
            '1004b04048e001000000', # TBD
            '1004b12048d000000000',
            '1004b1204848D3044110']
            for i, seq in enumerate(top_sequences, 1):
                byte_data = bytes.fromhex(seq)
                ser.write(byte_data)
                # 根据前缀决定是否读取响应
                if seq.startswith('1104'):
                    time.sleep(0.2)
                    response = ser.read_all()
                    while response.hex()!=11030000:
                        response = ser.read_all()
                        print(f"接收数据: {response.hex()}")
                    if response:
                        print(f"接收数据: {response.hex()}")
                    else:
                        print("[警告] 未收到响应数据")

    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def soc_init( port='COM21', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        flag=0
        top_sequences=["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                       "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"]
        for i, seq in enumerate(top_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                    big_endian_value = int.from_bytes(response[:4], byteorder='big')  # 大端解析
                    little_endian_bytes = big_endian_value.to_bytes(4, byteorder='little')  # 转为小端字节
                    print(f"接收数据: {little_endian_bytes.hex()}")
                else:
                    print("[警告] 未收到响应数据")
                    print("soc init failed")
                    flag = 1
                    break

        if(flag==0):
            print("===========soc init success==========")
    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
        exit()
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def phy_init(device):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0xf01e)
    print("0xf01e:", value)
    if (value == "0x4967"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()


def configure_from_txt(txt_file, device, lane):
    """
    从TXT文件读取并配置寄存器

    参数:
        txt_file: TXT文件路径
        device: 设备对象
        lane: 通道号
    """
    with open(txt_file, 'r') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()

        # 跳过空行和注释
        if not line or line.startswith('//') or line.startswith('#'):
            continue

        # 处理延迟
        if line.startswith('delay'):
            parts = line.split()
            if len(parts) > 1:
                delay_ms = int(parts[1])
                time.sleep(delay_ms / 1000.0)
            continue

        # 处理寄存器配置 (格式: 地址,值)
        if ',' in line:
            try:
                # 分割地址和值
                addr_str, value_str = line.split(',')

                # 转换为十六进制整数
                addr = int(addr_str.strip(), 16)
                value = int(value_str.strip(), 16)

                # 写入寄存器
                write_serdes_lane_register(device, lane, addr, value)
                time.sleep(0.005)
                value0=read_serdes_register(device, addr)
                print(f"addr=0x{addr:04X},{value0}")
                print(f"配置: lane={lane}, addr=0x{addr:04X}, value=0x{value:04X}")


            except ValueError as e:
                print(f"解析行失败 '{line}': {e}")



def notify(message, title="提示", timeout=5):
    """系统通知，自动消失"""
    notification.notify(
        title=title,
        message=message,
        timeout=timeout,  # 秒
        app_name="测试系统"
    )



if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    com_c = get_com_port_for_channel("C")
    ser=uartm_init(port=com_c, baudrate=115200)
    # 测试之前先reset
    phy_reset = [# phy_reset
                "1004b04048ccffffbc03", "1104b04048cc",]
    write_uartm(phy_reset,ser)
    soc_init = ["1004b04048e0040f7F00", "1104b04048e0",
                "1004b04048d000000000", "1104b04048d0",
                "1004b04048ccFFFFBC2b", "1104b04048CC"]
    write_uartm(soc_init,ser)
    ser.close()
    FTDI_CABLE_SERDES = b"A"  #
    device1 = setup_device(FTDI_CABLE_SERDES)
    phy_init(device1)
    txt_file = 'P5_TEST_TXT/phy_init.txt'  # ss的lowband使用8G_all_inlp_prbs7.txt
    configure_from_txt(txt_file, device1, 0)
    test(device1)
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    close_serdes(device1)
    print(f"总耗时: {duration}")
    notify("测试完成!!！", timeout=3)
    play_sound_test_complete()