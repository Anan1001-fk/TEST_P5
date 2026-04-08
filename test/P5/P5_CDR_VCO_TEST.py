import datetime
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import serial
import time
from instrument import uxr0134a
import openpyxl
import winsound

cfg_mode = "high_band"
code_mode = "XT"  # 粗调
def test():
   wb=openpyxl.load_workbook("P5_CDR_OPENLOOP_VCO.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   address = "TCPIP0::169.254.167.79::inst0::INSTR"
   uxr=uxr0134a.Uxr0134a(address)
   channal = 1
   chip_num = 'TT01'
   temp_volt = "NTNV"
   lane=0
   datarate = "32Gbps"
   pattern = "clk"
   if code_mode=="CT":
       for code in range(0, 1024):
           # 测粗调之前先把细调的code固定
           write_serdes_lane_register(device1, lane, 0xe050, 0x4)
           # 遍历粗调code
           # freq_tune over_en
           write_serdes_bit_reg(device1, lane, 0xe0a7, 11, 11, 1)
           # freq_tune
           write_serdes_bit_reg(device1, lane, 0xe0a7, 1, 10, code)
           # freq_tune_clk
           write_serdes_bit_reg(device1, lane, 0xe0a7, 0, 0, 1)
           time.sleep(0.1)
           if cfg_mode == "CT" & code_mode == "low_band":
               write_serdes_lane_register(device1, 0, 0xc018, 0xc00a)
               time.sleep(0.1)
               write_serdes_bit_reg(device1, 0, 0xc03f, 4, 4, 0)
               time.sleep(0.01)
               write_serdes_bit_reg(device1, 0, 0xc03f, 4, 4, 1)
           else:
               continue
           time.sleep(0.01)
           uxr.set_uxr_state("RUN")
           time.sleep(0.1)
           uxr.auto_verticalscale(channal)
           time.sleep(0.1)
           uxr.clear_display()
           time.sleep(0.1)
           uxr.set_uxr_state("stop")
           freq = uxr.get_meas_val(channal, "FREQuency", "MEAN")
           freq = round(float(freq) * 1e-6, 4)
           temp_list = [chip_num, temp_volt, lane, datarate, pattern, cfg_mode, code_mode, code, freq]
           print(f"当前code是{code},频率是{freq}")
           for i in range(len(temp_list)):
               ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               wb.save("P5_CDR_OPENLOOP_VCO.xlsx")
           last_row = last_row + 1
           temp_list.clear()
   else:
       for code in range(0, 1024):
           write_serdes_bit_reg(device1, lane, 0xe050, 4,13,code)
           time.sleep(0.01)
           uxr.set_uxr_state("RUN")
           time.sleep(0.1)
           uxr.auto_verticalscale(channal)
           time.sleep(0.1)
           uxr.clear_display()
           time.sleep(0.1)
           uxr.set_uxr_state("stop")
           freq = uxr.get_meas_val(channal, "FREQuency", "MEAN")
           freq = round(float(freq) * 1e-6, 4)
           temp_list = [chip_num, temp_volt, lane, datarate, pattern, cfg_mode, code_mode, code, freq]
           print(f"当前code是{code},频率是{freq}")
           for i in range(len(temp_list)):
               ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               wb.save("P5_CDR_OPENLOOP_VCO.xlsx")
           last_row = last_row + 1
           temp_list.clear()


def calculate_address(addr):
    """计算转换后的地址"""
    return 0xB0400000 + (0x1 << 19) + addr * 4
def initial_top(device):
    print("==============初始化top================")
    print("Verifying top...")
    init_top(device)

    '''
    [16:24:06.690] 读取成功 -> 地址: 0xb1204800, 数据: 0x20240821
    [16:24:06.691] 读取成功 -> 地址: 0xb1204804, 数据: 0x22334455
    [16:24:06.691] 读取成功 -> 地址: 0xb12048ac, 数据: 0x00000003
    [16:24:06.693] 读取成功 -> 地址: 0xb1204818, 数据: 0x00000001
    [16:24:06.693] 读取成功 -> 地址: 0xb1204848, 数据: 0x134d04d3
    [16:24:06.693] 读取成功 -> 地址: 0xb1204824, 数据: 0x02000000
    '''

    write_top_register(device, 0xb12048ac, 0x3)
    write_top_register(device, 0xb1204818, 0x1)
    write_top_register(device, 0xb1204848, 0x134d04d3)
    write_top_register(device, 0xb1204824, 0x2000000)

    print("0xb1204800:", read_top_register(device, 0xb1204800).hex())
    print("0xb1204804:", read_top_register(device, 0xb1204804).hex())
    print("0xb12048ac:", read_top_register(device, 0xb12048ac).hex())
    print("0xb1204818:", read_top_register(device, 0xb1204818).hex())
    print("0xb1204848:", read_top_register(device, 0xb1204848).hex())
    print("0xb1204824:", read_top_register(device, 0xb1204824).hex())

    print("==============top初始化成功================")
    time.sleep(0.1)
    print("==============*************************================")
def initial_serdes(device1):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0xf01e).hex()
    print("0xf01e:", value)
    if (value == "5983"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()
def configure_from_txt(txt_file, device):
    """
    从TXT文件读取并配置寄存器

    参数:
        txt_file: TXT文件路径
        device: 设备对象
        lane: 通道号
    """
    with open(txt_file, 'r') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()

        # 跳过空行和注释
        if not line or line.startswith('//') or line.startswith('#'):
            continue

        # 处理延迟
        if line.startswith('delay'):
            parts = line.split()
            if len(parts) > 1:
                delay_ms = int(parts[1])
                time.sleep(delay_ms / 1000.0)
            continue

        # 处理寄存器配置 (格式: 地址,值)
        if ',' in line:
            try:
                # 分割地址和值
                addr_str, value_str = line.split(',')

                # 转换为十六进制整数
                addr = int(addr_str.strip(), 16)
                value = int(value_str.strip(), 16)

                # 写入寄存器
                write_serdes_register(device, addr, value)
                time.sleep(0.005)
                value0=read_serdes_register(device,0, addr)
                print(f"addr=0x{addr:04X},{value0}")
                print(f"addr=0x{addr:04X}, value=0x{value:04X}")


            except ValueError as e:
                print(f"解析行失败 '{line}': {e}")



def soc_init( port='COM21', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        flag=0
        top_sequences=["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                       "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"]
        for i, seq in enumerate(top_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                    big_endian_value = int.from_bytes(response[:4], byteorder='big')  # 大端解析
                    little_endian_bytes = big_endian_value.to_bytes(4, byteorder='little')  # 转为小端字节
                    print(f"接收数据: {little_endian_bytes.hex()}")
                else:
                    print("[警告] 未收到响应数据")
                    print("soc init failed")
                    flag = 1
                    break

        if(flag==0):
            print("===========soc init success==========")
    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
        exit()
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def phy_init(device1):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1,0, 0xf01e)
    print("0xf01e:", value)
    if (value == "0x4967"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()

def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")


if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    soc_init(port='COM25', baudrate=115200)
    FTDI_CABLE_SERDES = b"A"  #
    device1 = setup_device(FTDI_CABLE_SERDES)
    phy_init(device1)
    txt_file = 'PCIE_32_QPLL0_direct_gui_contin_off.txt'  # 初始化到32G
    configure_from_txt(txt_file, device1)
    time.sleep(0.1)
    if cfg_mode=="CT":
        txt_file = 'RX_CDR_open_Kvco_frequency_range.txt'  # 配成开环
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        txt_file = 'RX2TXlpbk_clk_only.txt'  # 配成clk，送到tx
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
    else:
        txt_file = 'RX_CDR_open_Kvco_frequency_range_xt.txt'  # 配成开环
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        txt_file = 'RX2TXlpbk_clk_only_xt.txt'  # 配成clk，送到tx
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
    test()
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    close_serdes(device1)
    print(f"总耗时: {duration}")
    play_sound_test_complete()

