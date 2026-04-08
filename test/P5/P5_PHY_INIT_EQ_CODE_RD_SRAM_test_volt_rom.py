import openpyxl
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import datetime

from instrument import DM3058E
from instrument.DM3058E import Dm3058e


def save_register_data_to_excel(txt_file, device, lane, external_times, excel_file="DC.xlsx"):
    """
    读取txt文件中的寄存器信息，测量一次并保存到Excel，追加一行。
    external_times: 外部循环次数，用于填充times列。
    """
    # 解析寄存器信息（复用之前的 parse_register_txt 函数）
    register_info = parse_register_txt(txt_file, lane)
    if not register_info:
        print("没有找到任何寄存器信息")
        return

    # 检查文件是否存在，决定是否写入表头
    file_exists = os.path.isfile(excel_file)
    if file_exists:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.worksheets[0]
        last_row = ws.max_row+1
        if external_times==0:
            # 写入表头（寄存器名称）
            ws.cell(row=last_row, column=1, value="times")
            for col, info in enumerate(register_info, start=2):
                ws.cell(row=last_row, column=col, value=info['comment'])
            last_row += 1

            # 写入地址行
            ws.cell(row=last_row, column=1, value="")
            for col, info in enumerate(register_info, start=2):
                addr_str = f"0x{info['actual_addr']:04X}"
                ws.cell(row=last_row, column=col, value=addr_str)
            last_row += 1

            # 写入位段行
            ws.cell(row=last_row, column=1, value="")
            for col, info in enumerate(register_info, start=2):
                bits_str = f"[{info['high_bit']}:{info['low_bit']}]"
                ws.cell(row=last_row, column=col, value=bits_str)
            last_row += 1


        # 读取当前值并写入一行数据
        values = []
        for info in register_info:
            raw = read_serdes_register(device, info['actual_addr'])
            raw = int(raw, 16)
            width = info['high_bit'] - info['low_bit'] + 1
            mask = (1 << width) - 1
            bit_val = (raw >> info['low_bit']) & mask
            values.append(bit_val)

        ws.cell(row=last_row, column=1, value=external_times)
        for col, val in enumerate(values, start=2):
            ws.cell(row=last_row, column=col, value=f"0x{val:04X}")
        wb.save(excel_file)
        print(f"数据已追加到 {excel_file}，外部循环次数: {external_times}")

def soc_init( port='COM21', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        flag=0
        top_sequences=["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                       "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"]
        for i, seq in enumerate(top_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                    big_endian_value = int.from_bytes(response[:4], byteorder='big')  # 大端解析
                    little_endian_bytes = big_endian_value.to_bytes(4, byteorder='little')  # 转为小端字节
                    print(f"接收数据: {little_endian_bytes.hex()}")
                else:
                    print("[警告] 未收到响应数据")
                    print("soc init failed")
                    flag = 1
                    break

        if(flag==0):
            print("===========soc init success==========")
    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
        exit()
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")
# 使用示例（需先定义 read_serdes_register 函数）
if __name__ == "__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    full_lane = False
    lane = 0  # 逻辑通道号
    com_c=get_com_port_for_channel("C")
    port=com_c
    baudrate=115200

    ser=uartm_init(port=port, baudrate=baudrate)
    for times in range(10):
        # part1
        init = ["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"
                # phy_reset
                "1004b04048ccffffbc03", "1104b04048cc",
                # release_phy_reset
                 "1004b04048ccffffbc2b","1104b04048cc" ,]
        write_uartm(init,ser)
        FTDI_CABLE_SERDES = b"A"  #
        device1 = setup_device(FTDI_CABLE_SERDES)
        phy_init(device1)
        txt_file = 'P5_TEST_TXT/powermode_to_L0.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        # write_serdes_bit_reg(device1, lane, 0xe010, 2, 10, 0x10)
        #write_serdes_bit_reg(device1,lane,0xe13f,0,14,0x1f)
        #write_serdes_bit_reg(device1, lane, 0xe048, 0, 7, 0xff)
        #write_serdes_bit_reg(device1, lane, 0xe06c, 0, 2, 0)
        write_serdes_lane_register(device1, lane, 0xe058, 0x8C00)
        time.sleep(1)
        write_serdes_lane_register(device1, lane, 0xe058, 0x8C00)
        print("清掉误码后，等待1s")
        time.sleep(1)
        err_count = read_serdes_register(device1, 0xe057)
        time.sleep(0.1)
        if full_lane == True:
            for lane in range(4):
                txt_file = "P5_TEST_TXT/phy_init_cal_eq_code_rd_final.txt"
                save_register_data_to_excel(txt_file, device1, lane, external_times=times, excel_file="read_eq_code_volt.xlsx")
        else:
            txt_file = "P5_TEST_TXT/phy_init_cal_eq_code_rd_final.txt"
            save_register_data_to_excel(txt_file, device1, lane, external_times=times, excel_file="read_eq_code_volt.xlsx")
        dm1 = DM3058E.Dm3058e(address="USB0::0x1AB1::0x09C4::DM3L273100483::INSTR")
        dm2 = DM3058E.Dm3058e(address="USB0::0x1AB1::0x09C4::DM3R261200875::INSTR")
        result = []
        wb = openpyxl.load_workbook("read_eq_code_volt.xlsx")
        ws = wb.worksheets[0]
        last_row = ws.max_row
        txt_file = 'P5_TEST_TXT/RXAFE_ofst_cal_att_mear.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        att_n=dm2.get_voltage()
        print("att_n电压是",att_n)
        att_p = dm1.get_voltage()
        print("att_p电压是", att_p)
        result.append("att_n")
        result.append(att_n)
        result.append("att_p")
        result.append(att_p)
        txt_file = 'P5_TEST_TXT/RXAFE_ofst_cal_buf_mear.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        buf_n = dm2.get_voltage()
        print("buf_n电压是", buf_n)
        buf_p = dm1.get_voltage()
        print("buf_p电压是", buf_p)
        result.append("buf_n")
        result.append(buf_n)
        result.append("buf_p")
        result.append(buf_p)
        txt_file = 'P5_TEST_TXT/RXAFE_ofst_cal_ctle_mear.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        ctle_n = dm2.get_voltage()
        print("ctle_n电压是", ctle_n)
        ctle_p = dm1.get_voltage()
        print("ctle_p电压是", ctle_p)
        result.append("ctle_n")
        result.append(ctle_n)
        result.append("ctle_p")
        result.append(ctle_p)
        txt_file = 'P5_TEST_TXT/RXAFE_ofst_cal_vga_mear.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        vga_n = dm2.get_voltage()
        print("vga_n电压是", vga_n)
        vga_p = dm1.get_voltage()
        print("vga_p电压是", vga_p)
        result.append("vga_n")
        result.append(vga_n)
        result.append("vga_p")
        result.append(vga_p)
        for i in range(len(result)):
            ws.cell(row=last_row, column=i+71, value=result[i])
        wb.save("read_eq_code_volt.xlsx")
        result.clear()
        close_serdes(device1)



