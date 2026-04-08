import winsound
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import openpyxl
import datetime
import serial
import time
import openpyxl
from instrument import uxr0134a

def test():
   wb=openpyxl.load_workbook("pll_vco_band.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   address = "TCPIP0::169.254.161.219::inst0::INSTR"
   uxr=uxr0134a.Uxr0134a(address)
   channal=1
   chip_num="TT01"
   temp_volt="NTNV"
   txt_file="ate_nearpma_PCIE_32GBPS_QPLL0_direct_r260306.txt"
   configure_from_txt(txt_file,device1)
   # 断开Qpll环路
   write_serdes_lane_register(device1,0,0xf09f,0xbd8b)
   # 设置固定的vctrl
   write_serdes_lane_register(device1,0,0xf05b,0x8000)
   for fine_tune in range(4):
       write_serdes_bit_reg(device1, 0, 0xf016, 9, 10, fine_tune)
       write_serdes_bit_reg(device1, 0, 0xf017, 9, 10, fine_tune)
       for coarse_tune in range(10):
           write_serdes_bit_reg(device1,0,0xf0b7,0,6,coarse_tune)
           uxr.set_uxr_state("RUN")
           time.sleep(0.1)
           uxr.auto_verticalscale(channal)
           time.sleep(0.1)
           uxr.clear_display()
           time.sleep(0.1)
           uxr.set_uxr_state("stop")
           freq = uxr.get_meas_val(channal, "FREQuency", "MEAN")
           freq = round(float(freq) * 1e-9, 3)
           print(f"目前测试的finetune的code是{fine_tune}，coarsetune的code是{coarse_tune}，此时频率是{freq}")
           temp_list = [chip_num, temp_volt, fine_tune, coarse_tune, freq]
           for i in range(len(temp_list)):
               ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               wb.save("pll_vco_band.xlsx")
           last_row = last_row + 1
           temp_list.clear()




def soc_init( port='COM21', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        flag=0
        top_sequences=["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                       "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"]
        for i, seq in enumerate(top_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                    big_endian_value = int.from_bytes(response[:4], byteorder='big')  # 大端解析
                    little_endian_bytes = big_endian_value.to_bytes(4, byteorder='little')  # 转为小端字节
                    print(f"接收数据: {little_endian_bytes.hex()}")
                else:
                    print("[警告] 未收到响应数据")
                    print("soc init failed")
                    flag = 1
                    break

        if(flag==0):
            print("===========soc init success==========")
    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def phy_init(device):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0xf01e)
    print("0xf01e:", value)
    if (value == "0x4967"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()


def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")


def configure_from_txt(txt_file, device):
    """
    从TXT文件读取并配置寄存器

    参数:
        txt_file: TXT文件路径
        device: 设备对象
        lane: 通道号
    """
    with open(txt_file, 'r') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()

        # 跳过空行和注释
        if not line or line.startswith('//') or line.startswith('#'):
            continue

        # 处理延迟
        if line.startswith('delay'):
            parts = line.split()
            if len(parts) > 1:
                delay_ms = int(parts[1])
                time.sleep(delay_ms / 1000.0)
            continue

        # 处理寄存器配置 (格式: 地址,值)
        if ',' in line:
            try:
                # 分割地址和值
                addr_str, value_str = line.split(',')

                # 转换为十六进制整数
                addr = int(addr_str.strip(), 16)
                value = int(value_str.strip(), 16)

                # 写入寄存器
                write_serdes_register(device, addr, value)
                time.sleep(0.005)
                value0=read_serdes_register(device, addr)
                print(f"addr=0x{addr:04X},{value0}")
                print(f"addr=0x{addr:04X}, value=0x{value:04X}")


            except ValueError as e:
                print(f"解析行失败 '{line}': {e}")
if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    soc_init(port='COM25', baudrate=115200)
    FTDI_CABLE_SERDES = b"A"  #
    device1 = setup_device(FTDI_CABLE_SERDES)
    phy_init(device1)
    test()
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    close_serdes(device1)
    print(f"总耗时: {duration}")
    play_sound_test_complete()