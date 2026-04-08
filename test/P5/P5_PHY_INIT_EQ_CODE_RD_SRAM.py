import openpyxl
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import datetime






def save_register_data_to_excel(txt_file, device, lane, external_times, excel_file="DC.xlsx"):
    """
    读取txt文件中的寄存器信息，测量一次并保存到Excel，追加一行。
    external_times: 外部循环次数，用于填充times列。
    """
    # 解析寄存器信息（复用之前的 parse_register_txt 函数）
    register_info = parse_register_txt(txt_file, lane)
    if not register_info:
        print("没有找到任何寄存器信息")
        return

    # 检查文件是否存在，决定是否写入表头
    file_exists = os.path.isfile(excel_file)
    if file_exists:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.worksheets[0]
        last_row = ws.max_row+1
        if external_times==0:
            # 写入表头（寄存器名称）
            ws.cell(row=last_row, column=1, value="times")
            for col, info in enumerate(register_info, start=2):
                ws.cell(row=last_row, column=col, value=info['comment'])
            last_row += 1

            # 写入地址行
            ws.cell(row=last_row, column=1, value="")
            for col, info in enumerate(register_info, start=2):
                addr_str = f"0x{info['actual_addr']:04X}"
                ws.cell(row=last_row, column=col, value=addr_str)
            last_row += 1

            # 写入位段行
            ws.cell(row=last_row, column=1, value="")
            for col, info in enumerate(register_info, start=2):
                bits_str = f"[{info['high_bit']}:{info['low_bit']}]"
                ws.cell(row=last_row, column=col, value=bits_str)
            last_row += 1


        # 读取当前值并写入一行数据
        values = []
        for info in register_info:
            raw = read_serdes_register(device, info['actual_addr'])
            raw = int(raw, 16)
            width = info['high_bit'] - info['low_bit'] + 1
            mask = (1 << width) - 1
            bit_val = (raw >> info['low_bit']) & mask
            values.append(bit_val)

        ws.cell(row=last_row, column=1, value=external_times)
        for col, val in enumerate(values, start=2):
            ws.cell(row=last_row, column=col, value=f"0x{val:04X}")
        wb.save(excel_file)
        print(f"数据已追加到 {excel_file}，外部循环次数: {external_times}")


# 使用示例（需先定义 read_serdes_register 函数）
if __name__ == "__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    full_lane = False
    lane = 0  # 逻辑通道号
    port="COM25"
    baudrate=115200
    ser=uartm_init(port=port, baudrate=baudrate)
    for times in range(5):
        # part1
        part1 = ["1004b04048ccffffbc03", "1104b04048cc",
                 "1004b062000410000000", "1104b0620004",
                 "1004b04048e0040f7f00", "1104b04048e0",
                 "1004b04048d000000000", "1104b04048d0",
                 # lane_reset
                 "1004b060000803000000", "1104b0600008",
                 # switch_sram
                 "1004b0620b1400000000", "1104b0620b14",
                # release_phy_reset
                 "1004b04048ccffffbc2b","1104b04048cc" ,
                 "1004b0620b1430000000","1104b0620b14"]
        write_uartm(part1,ser)
        # update
        update = ["1004b0620b1420000000", "1104b0620b14"]
        write_uartm(update,ser)
        FTDI_CABLE_SERDES = b"A"  #
        device1 = setup_device(FTDI_CABLE_SERDES)
        phy_init(device1)
        time.sleep(0.1)
        txt_file="eq4_target50_tap1wgt2_dfe_tap1to5_1.20260324.rom_enc_new.txt"
        write_32bit_data_from_txt(txt_file,device1,0,1,1)
        time.sleep(0.1)
        close_serdes(device1)
        # part2
        part2 = [# set phy0_sram_apb_sel=0,set ext_load_done =1
                 "1004b0620b1428000000", "1104b0620b14",
                # release lane_reset
                 "1004b060000802000000","1104b0600008",]
        write_uartm(part2,ser)
        device1 = setup_device(FTDI_CABLE_SERDES)
        phy_init(device1)
        txt_file = 'ate__pcie_32G_qpll0_casc_lpbk_ext_afeos_afeadapt_sram(2).txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        txt_file = 'cont_os_cal_final.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        write_serdes_bit_reg(device1, lane, 0xe010, 2, 10, 0x10)
        #write_serdes_bit_reg(device1,lane,0xe13f,0,14,0x1f)
        #write_serdes_bit_reg(device1, lane, 0xe048, 0, 7, 0xff)
        #write_serdes_bit_reg(device1, lane, 0xe06c, 0, 2, 0)
        txt_file = 'AFE_DFE_adapt_mode4.txt'  # 初始化到32G
        configure_from_txt(txt_file, device1)
        time.sleep(0.1)
        write_serdes_lane_register(device1, lane, 0xe058, 0x8C00)
        time.sleep(1)
        write_serdes_lane_register(device1, lane, 0xe058, 0x8C00)
        print("清掉误码后，等待1min")
        time.sleep(60)

        err_count = read_serdes_register(device1, 0xe057)
        time.sleep(0.1)
        if full_lane == True:
            for lane in range(4):
                txt_file = "phy_init_cal_eq_code_rd_final.txt"
                save_register_data_to_excel(txt_file, device1, lane, external_times=times, excel_file="read_eq_code.xlsx")
        else:
            txt_file = "phy_init_cal_eq_code_rd_final.txt"
            save_register_data_to_excel(txt_file, device1, lane, external_times=times, excel_file="read_eq_code.xlsx")
        close_serdes(device1)



