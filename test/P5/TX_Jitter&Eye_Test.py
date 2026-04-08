import winsound
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import datetime
import serial
import time
import openpyxl
from instrument import uxr0134a


def test():
   wb=openpyxl.load_workbook("P5_Jitter&Eye.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   address = "TCPIP0::169.254.161.219::inst0::INSTR"
   uxr=uxr0134a.Uxr0134a(address)
   channal=1
   chip_num="TT01"
   temp_volt="NTNV"
   lane=0
   index=lane
   datarate_list=[32,16]
   txt_file_list=["32Gbps.txt","16Gbps.txt"]
   pattern_list = {
       "clk":    0x9d54,
       "prbs7":  0x1d54,
       "prbs15": 0xad54,
       "prbs31": 0x8d54,}
   for datarate, txt_file in zip(datarate_list, txt_file_list):
       print(f"\n===== 开始测试速率 {datarate}Gbps =====")
       # 1. 配置速率基础
       configure_from_txt(txt_file, device1)
       time.sleep(0.2)  # 等待稳定

       # 2. 测试CLK
       print("--- 测试CLK ---")
       write_serdes_lane_register(device1, index, 0xe02d, pattern_list["clk"])
       time.sleep(0.05)
       main = read_serdes_bit_reg(device1,index, 0xe01c, 9, 14)
       pre = read_serdes_bit_reg(device1, index,0xe01b, 0, 5)
       post = read_serdes_bit_reg(device1,index, 0xe01b, 7, 12)
       uxr.set_uxr_state("RUN")
       uxr.set_clock_recovery_ojtf(method="SOPLL",data_rate=datarate*1e9,loop_bandwidth=datarate*1e9/1667)
       uxr.set_mem_depth()
       uxr.auto_verticalscale(channal)
       time.sleep(2)
       uxr.clear_display()
       time.sleep(5)
       uxr.set_uxr_state("stop")
       time.sleep(0.1)
       vpp = uxr.get_meas_val(channal, "VPP",meas_statistic="MEAN")
       vpp = vpp * 1000
       eyehight = uxr.get_eh_value_mv()
       eyewidth = uxr.get_ew_value_ps()
       jitter = uxr.get_tj()
       time.sleep(2)
       tj = jitter[0]
       tj = round(float(tj) * 1e12, 3)
       rj = jitter[1]
       rj = round(float(rj) * 1e12, 3)
       dj = jitter[2]
       dj = round(float(dj) * 1e12, 3)
       pj = jitter[4]
       pj = round(float(pj) * 1e12, 3)
       isi = jitter[7]
       isi = round(float(isi) * 1e12, 3)
       dcd = jitter[6]
       dcd = round(float(dcd) * 1e12, 3)
       temp_list = [chip_num, temp_volt, datarate, index, "clk", main, pre, post, vpp,
                    eyehight, eyewidth, tj, rj, dj, pj, isi,dcd]
       uxr.save_screen_image(
           f"C:\\FK\\{chip_num}_{temp_volt}_{datarate}Gbps_lane{lane}_clk_main={main}_pre={pre}_post={post}.png")  # default path:C:\Users\Public\Documents\Infiniium
       for i in range(len(temp_list)):
           ws.cell(row=last_row, column=i + 1, value=temp_list[i])
           wb.save("P5_Jitter&Eye.xlsx")
       last_row = last_row + 1
       temp_list.clear()

       # 3. PRBS7优化
       print("--- PRBS7优化 ---")
       # 先配置PRBS7
       write_serdes_lane_register(device1, index, 0xe02d, pattern_list["prbs7"])
       time.sleep(0.05)

       # 3.1 遍历post
       print("遍历post...")
       best_post = None
       best_isi_post = float('inf')
       for post_val in range(0, 22, 2):  # 2,4,...,30
           # 设置post
           write_serdes_bit_reg(device1, index, 0xe01b, 7, 12, post_val)
           time.sleep(0.01)  # 等待参数生效
           main = read_serdes_bit_reg(device1, index,0xe01c, 9, 14)
           pre = read_serdes_bit_reg(device1, index,0xe01b, 0, 5)
           post = read_serdes_bit_reg(device1,index, 0xe01b, 7, 12)
           uxr.set_uxr_state("RUN")
           uxr.auto_verticalscale(channal)
           time.sleep(2)
           uxr.clear_display()
           time.sleep(5)
           uxr.set_uxr_state("stop")
           time.sleep(0.1)
           vpp = uxr.get_meas_val(channal, "VPP", meas_statistic="MEAN")
           vpp = vpp * 1000
           eyehight = uxr.get_eh_value_mv()
           eyewidth = uxr.get_ew_value_ps()
           jitter = uxr.get_tj()
           time.sleep(2)
           tj = jitter[0]
           tj = round(float(tj) * 1e12, 3)
           rj = jitter[1]
           rj = round(float(rj) * 1e12, 3)
           dj = jitter[2]
           dj = round(float(dj) * 1e12, 3)
           pj = jitter[4]
           pj = round(float(pj) * 1e12, 3)
           isi = jitter[7]
           isi = round(float(isi) * 1e12, 3)
           dcd = jitter[6]
           dcd = round(float(dcd) * 1e12, 3)
           print(f"  post={post_val}, ISI={isi}ps")
           if isi < best_isi_post:
               best_isi_post = isi
               best_post = post_val

           temp_list = [chip_num, temp_volt, datarate, index, "prbs7", main, pre, post, vpp,
                        eyehight, eyewidth, tj, rj, dj, pj, isi, dcd]
           uxr.save_screen_image(
               f"C:\\FK\\{chip_num}_{temp_volt}_{datarate}Gbps_lane{lane}_prbs7_main={main}_pre={pre}_post={post}.png")  # default path:C:\Users\Public\Documents\Infiniium
           for i in range(len(temp_list)):
               ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               wb.save("P5_Jitter&Eye.xlsx")
           last_row = last_row + 1
           temp_list.clear()
       # 固定最佳post
       write_serdes_bit_reg(device1, index, 0xe01b, 7, 12, best_post)
       print(f"最佳post={best_post}, ISI={best_isi_post}ps")

       # 3.2 遍历pre
       print("遍历pre...")
       best_pre = None
       best_isi_pre = float('inf')
       for pre_val in range(0, 16, 2):
           # 设置pre
           write_serdes_bit_reg(device1, index, 0xe01b, 0, 5, pre_val)
           time.sleep(0.01)
           main = read_serdes_bit_reg(device1,index,0xe01c, 9, 14)
           pre =  read_serdes_bit_reg(device1, index,0xe01b, 0, 5)
           post = read_serdes_bit_reg(device1,index, 0xe01b, 7, 12)
           uxr.set_uxr_state("RUN")
           uxr.auto_verticalscale(channal)
           time.sleep(2)
           uxr.clear_display()
           time.sleep(5)
           uxr.set_uxr_state("stop")
           time.sleep(0.1)
           vpp = uxr.get_meas_val(channal, "VPP", meas_statistic="MEAN")
           vpp = vpp * 1000
           eyehight = uxr.get_eh_value_mv()
           eyewidth = uxr.get_ew_value_ps()
           jitter = uxr.get_tj()
           time.sleep(2)
           tj = jitter[0]
           tj = round(float(tj) * 1e12, 3)
           rj = jitter[1]
           rj = round(float(rj) * 1e12, 3)
           dj = jitter[2]
           dj = round(float(dj) * 1e12, 3)
           pj = jitter[4]
           pj = round(float(pj) * 1e12, 3)
           isi = jitter[7]
           isi = round(float(isi) * 1e12, 3)
           dcd = jitter[6]
           dcd = round(float(dcd) * 1e12, 3)
           print(f"  pre={pre_val}, ISI={isi}ps")
           if isi < best_isi_pre:
               best_isi_pre = isi
               best_pre = pre_val

           temp_list = [chip_num, temp_volt, datarate, index, "prbs7", main, pre, post, vpp,
                        eyehight, eyewidth, tj, rj, dj, pj, isi, dcd]
           uxr.save_screen_image(
               f"C:\\FK\\{chip_num}_{temp_volt}_{datarate}Gbps_lane{lane}_prbs7_main={main}_pre={pre}_post={post}.png")  # default path:C:\Users\Public\Documents\Infiniium
           for i in range(len(temp_list)):
               ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               wb.save("P5_Jitter&Eye.xlsx")
           last_row = last_row + 1
           temp_list.clear()
       # 固定最佳pre
       write_serdes_bit_reg(device1, index, 0xe01b, 0, 5, best_pre)
       print(f"最佳pre={best_pre}, ISI={best_isi_pre}ps")

       # 4. 测试PRBS15
       print("--- 测试PRBS15 ---")
       uxr.set_mem_depth(2e6)
       write_serdes_lane_register(device1, index, 0xe02d, pattern_list["prbs15"])
       time.sleep(0.05)
       main = read_serdes_bit_reg(device1,index, 0xe01c, 9, 14)
       pre = read_serdes_bit_reg(device1,index, 0xe01b, 0, 5)
       post = read_serdes_bit_reg(device1,index, 0xe01b, 7, 12)
       uxr.set_uxr_state("RUN")
       uxr.auto_verticalscale(channal)
       time.sleep(2)
       uxr.clear_display()
       time.sleep(60)
       print(f"目前测试{datarate}Gbps的prbs15，等待1min收集jitter")
       uxr.set_uxr_state("stop")
       time.sleep(0.1)
       vpp = uxr.get_meas_val(channal, "VPP", meas_statistic="MEAN")
       vpp = vpp * 1000
       eyehight = uxr.get_eh_value_mv()
       eyewidth = uxr.get_ew_value_ps()
       jitter = uxr.get_tj()
       time.sleep(2)
       tj = jitter[0]
       tj = round(float(tj) * 1e12, 3)
       rj = jitter[1]
       rj = round(float(rj) * 1e12, 3)
       dj = jitter[2]
       dj = round(float(dj) * 1e12, 3)
       pj = jitter[4]
       pj = round(float(pj) * 1e12, 3)
       isi = jitter[7]
       isi = round(float(isi) * 1e12, 3)
       dcd = jitter[6]
       dcd = round(float(dcd) * 1e12, 3)
       temp_list = [chip_num, temp_volt, datarate, index, "prbs15", main, pre, post, vpp, eyehight, eyewidth, tj,
                    rj, dj, pj, isi,
                    dcd]
       uxr.save_screen_image(
           f"C:\\FK\\{chip_num}_{temp_volt}_{datarate}Gbps_lane{lane}_prbs15_main={main}_pre={pre}_post={post}.png")  # default path:C:\Users\Public\Documents\Infiniium
       for i in range(len(temp_list)):
           ws.cell(row=last_row, column=i + 1, value=temp_list[i])
           wb.save("P5_Jitter&Eye.xlsx")
       last_row = last_row + 1
       temp_list.clear()
       uxr.set_mem_depth(1e6)

       # 5. 测试PRBS31
       print("--- 测试PRBS31 ---")
       uxr.set_jitter_pattern_arbitrary()
       write_serdes_lane_register(device1, index, 0xe02d, pattern_list["prbs31"])
       time.sleep(0.05)
       main = read_serdes_bit_reg(device1,index, 0xe01c, 9, 14)
       pre = read_serdes_bit_reg(device1,index, 0xe01b, 0, 5)
       post = read_serdes_bit_reg(device1,index, 0xe01b, 7, 12)
       uxr.set_uxr_state("RUN")
       uxr.auto_verticalscale(channal)
       time.sleep(2)
       uxr.clear_display()
       time.sleep(5)
       uxr.set_uxr_state("stop")
       time.sleep(0.1)
       vpp = uxr.get_meas_val(channal, "VPP",meas_statistic="MEAN")
       vpp = vpp * 1000
       eyehight = uxr.get_eh_value_mv()
       eyewidth = uxr.get_ew_value_ps()
       jitter = uxr.get_tj()
       time.sleep(2)
       tj = jitter[0]
       tj = round(float(tj) * 1e12, 3)
       rj = jitter[1]
       rj = round(float(rj) * 1e12, 3)
       dj = jitter[2]
       dj = round(float(dj) * 1e12, 3)
       pj = jitter[4]
       pj = round(float(pj) * 1e12, 3)
       isi = jitter[7]
       isi = round(float(isi) * 1e12, 3)
       dcd = jitter[6]
       dcd = round(float(dcd) * 1e12, 3)
       temp_list = [chip_num, temp_volt, datarate, index, "prbs31", main, pre, post, vpp, eyehight, eyewidth, tj,
                    rj, dj, pj, isi,
                    dcd]
       uxr.save_screen_image(
           f"C:\\FK\\{chip_num}_{temp_volt}_{datarate}Gbps_lane{lane}_prbs31_main={main}_pre={pre}_post={post}.png")  # default path:C:\Users\Public\Documents\Infiniium
       for i in range(len(temp_list)):
           ws.cell(row=last_row, column=i + 1, value=temp_list[i])
           wb.save("P5_Jitter&Eye.xlsx")
       last_row = last_row + 1
       temp_list.clear()
       uxr.set_jitter_pattern_periodic()
       time.sleep(0.2)




def soc_init( port='COM21', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        flag=0
        top_sequences=["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                       "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"]
        for i, seq in enumerate(top_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                    big_endian_value = int.from_bytes(response[:4], byteorder='big')  # 大端解析
                    little_endian_bytes = big_endian_value.to_bytes(4, byteorder='little')  # 转为小端字节
                    print(f"接收数据: {little_endian_bytes.hex()}")
                else:
                    print("[警告] 未收到响应数据")
                    print("soc init failed")
                    flag = 1
                    break

        if(flag==0):
            print("===========soc init success==========")
    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def phy_init(device):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0xf01e)
    print("0xf01e:", value)
    if (value == "0x4967"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()


def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")



if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    com_c=get_com_port_for_channel("C")
    soc_init(port=com_c, baudrate=115200)
    FTDI_CABLE_SERDES = b"A"  #
    device1 = setup_device(FTDI_CABLE_SERDES)
    phy_init(device1)
    test()
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    close_serdes(device1)
    print(f"总耗时: {duration}")
    play_sound_test_complete()