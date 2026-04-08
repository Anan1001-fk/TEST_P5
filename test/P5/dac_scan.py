import winsound
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import openpyxl
from instrument import DM3058E
import datetime
import serial


def test(device1,lane):
   wb=openpyxl.load_workbook("dac_scan.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   chip_num = "TT01"
   temp_volt = "NTNV"
   value_list = []
   DM3058E.init()
   DM3058E.reset()
   time.sleep(1)
   temp_list = [chip_num, temp_volt]
   if txt_file.startswith("tap"):
       for code in range(0, 512):
            print("目前测试lane", int(lane))
            name_without_ext = os.path.splitext(txt_file)[0]
            value_list.append(int(lane))
            value_list.append(name_without_ext)
            value_list.append(code)
            write_serdes_lane_register(device1,lane,0xe0a2,0x8000)
            write_serdes_bit_reg(device1,lane,0xe0a4,0,8,code)
            write_serdes_lane_register(device1, lane, 0xe0d7, 0x8000)
            volt = DM3058E.get_voltage()
            time.sleep(0.1)
            print("电压是", volt, "v")
            value_list.append(volt)
            time.sleep(0.1)
            for i in range(len(temp_list)):
                ws.cell(row=last_row, column=i + 1, value=temp_list[i])
            for i in range(len(value_list)):
                ws.cell(row=last_row, column=i + 3, value=value_list[i])
            wb.save("dac_scan.xlsx")
            save_test_data_to_txt("dac_scan", temp_list, value_list)
            last_row = last_row + 1
            value_list.clear()
   else:
       for code in range(0, 256):
           print("目前测试lane", int(lane))
           value_list.append(int(lane))
           write_serdes_lane_register(device1, lane, 0xe0a2, 0x8000)
           write_serdes_bit_reg(device1, lane, 0xe0a4, 0, 7, code)
           write_serdes_lane_register(device1, lane, 0xe0d7, 0x8000)
           volt = DM3058E.get_voltage()
           time.sleep(0.1)
           print("电压是", volt, "v")
           value_list.append(volt)
           time.sleep(0.1)
           for i in range(len(temp_list)):
               ws.cell(row=last_row, column=i + 1, value=temp_list[i])
           for i in range(len(value_list)):
               ws.cell(row=last_row, column=i + 4, value=value_list[i])
           wb.save("P5_DC.xlsx")
           save_test_data_to_txt("DC", temp_list, value_list)
           last_row = last_row + 1
           value_list.clear()









def save_test_data_to_txt(filename, temp_list, value_list, timestamp=None):
    """
    保存测试数据到TXT文件

    参数:
        filename: 文件名
        temp_list: [芯片编号, 温度/电压, 极性]
        value_list: [测试通道, 测试程序, 电压值]
        timestamp: 时间戳（可选）
    """
    import datetime

    # 如果没有提供时间戳，生成当前时间
    if timestamp is None:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 合并所有数据
    all_data = temp_list + value_list + [timestamp]

    # 转换为字符串
    data_line = ','.join(str(item) for item in all_data)

    # 检查文件是否存在，决定是否添加表头
    if not os.path.exists(filename):
        # 文件不存在，创建并添加表头
        header = "芯片编号,温度/电压,极性,测试通道,测试程序,电压值,时间戳"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(header + '\n')
            f.write(data_line + '\n')
    else:
        # 文件存在，追加数据
        with open(filename, 'a', encoding='utf-8') as f:
            f.write(data_line + '\n')

    print(f"数据已保存到: {filename}")
    return True

def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")

def conditional_serial_communication(flag,code,port='COM5', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        # 这个是为了让cfg top&phy的寄存器只走一次
        if flag==1:
            # cfg top通路
            top_sequences = [
            '1104b040485000000000',
            '1004b04048e001000000', # TBD
            '1004b12048d000000000',
            '1004b1204848D3044110']
            for i, seq in enumerate(top_sequences, 1):
                byte_data = bytes.fromhex(seq)
                ser.write(byte_data)
                # 根据前缀决定是否读取响应
                if seq.startswith('1104'):
                    time.sleep(0.2)
                    response = ser.read_all()
                    while response.hex()!=11030000:
                        response = ser.read_all()
                        print(f"接收数据: {response.hex()}")
                    if response:
                        print(f"接收数据: {response.hex()}")
                    else:
                        print("[警告] 未收到响应数据")

    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def soc_init( port='COM21', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        flag=0
        top_sequences=["1104b0404850","1004b04048e0040f7F00","1104b04048e0",
                       "1004b04048d000000000","1104b04048d0","1004b04048ccFFFFBC2b","1104b04048CC"]
        for i, seq in enumerate(top_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                    big_endian_value = int.from_bytes(response[:4], byteorder='big')  # 大端解析
                    little_endian_bytes = big_endian_value.to_bytes(4, byteorder='little')  # 转为小端字节
                    print(f"接收数据: {little_endian_bytes.hex()}")
                else:
                    print("[警告] 未收到响应数据")
                    print("soc init failed")
                    flag = 1
                    break

        if(flag==0):
            print("===========soc init success==========")
    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
        exit()
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")


def phy_init(device1):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0,0xf01e)
    print("0xf01e:", value)
    if (value == "0x4967"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()


def configure_from_txt(txt_file, device, lane):
    """
    从TXT文件读取并配置寄存器

    参数:
        txt_file: TXT文件路径
        device: 设备对象
        lane: 通道号
    """
    with open(txt_file, 'r') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()

        # 跳过空行和注释
        if not line or line.startswith('//') or line.startswith('#'):
            continue

        # 处理延迟
        if line.startswith('delay'):
            parts = line.split()
            if len(parts) > 1:
                delay_ms = int(parts[1])
                time.sleep(delay_ms / 1000.0)
            continue

        # 处理寄存器配置 (格式: 地址,值)
        if ',' in line:
            try:
                # 分割地址和值
                addr_str, value_str = line.split(',')

                # 转换为十六进制整数
                addr = int(addr_str.strip(), 16)
                value = int(value_str.strip(), 16)

                # 写入寄存器
                write_serdes_lane_register(device, lane, addr, value)
                time.sleep(0.005)
                value0=read_serdes_register(device,lane, addr)
                print(f"addr=0x{addr:04X},{value0}")
                print(f"配置: lane={lane}, addr=0x{addr:04X}, value=0x{value:04X}")


            except ValueError as e:
                print(f"解析行失败 '{line}': {e}")

if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    soc_init(port='COM25', baudrate=115200)
    FTDI_CABLE_SERDES = b"A"  #
    device1 = setup_device(FTDI_CABLE_SERDES)
    phy_init(device1)
    lane=0
    txt_file = 'PCIE_32_QPLL0_direct_gui_contin_off.txt'  # 测试不同的dac需要更换不同的dac脚本
    configure_from_txt(txt_file, device1, lane)
    time.sleep(0.1)
    txt_file = 'P5_TEST_TXT/RXAFE_ofst_cal_att_mear.txt'  # 测试不同的dac需要更换不同的dac脚本
    configure_from_txt(txt_file, device1, lane)
    test(device1,lane)
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    close_serdes(device1)
    print(f"总耗时: {duration}")
    play_sound_test_complete()