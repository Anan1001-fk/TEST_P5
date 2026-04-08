import datetime
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import serial
import time
from instrument import uxr0134a
import openpyxl
import tkinter as tk
from tkinter import messagebox

def test():
   wb=openpyxl.load_workbook("11.xlsx")
   ws=wb.worksheets[3]
   last_row=ws.max_row+1
   address = "TCPIP0::169.254.202.222::inst0::INSTR"
   uxr=uxr0134a.Uxr0134a(address)
   flag=1
   channal = 2
   chip_num = 'SS10'
   temp_volt = "NTHV"
   datarate = "16Gbps"
   pattern = "clk"
   cfg_mode="low_band"
   code_mode="CT" # 粗调
   cfg_ct=['1004b04b82d400800000','1004b04b82d400820000','1004b04b82d400840000','1004b04b82d400860000','1004b04b82d400880000'
        ,'1004b04b82d4008A0000','1004b04b82d4008C0000','1004b04b82d4008E0000','1004b04b82d400900000','1004b04b82d400920000'
        ,'1004b04b82d400940000','1004b04b82d400960000','1004b04b82d400980000','1004b04b82d4009A0000','1004b04b82d4009C0000'
        ,'1004b04b82d4009E0000','1004b04b82d4F89F0000'] # 粗调code

   cfg_xt=['1004b04b811800000000','1004b04b811800040000','1004b04b811800080000','1004b04b8118000c0000','1004b04b811800100000'
        ,'1004b04b811800140000','1004b04b811800180000','1004b04b8118001c0000','1004b04b811800200000','1004b04b811800240000'
        ,'1004b04b811800280000','1004b04b8118002c0000','1004b04b811800300000','1004b04b811800340000','1004b04b811800380000'
        ,'1004b04b8118003c0000','1004b04b8118f03f0000']    #细调code 细调之前得先确认是high band还是low band
   if code_mode == "CT":
       cfg=cfg_ct
   else:
       if cfg_mode == "high_band":
           # 这个code根据粗调扫描对应到8Ghz来调节
           code ='1004b04b82d400900000'
       else:
           # 这个code根据粗调扫描对应到4Ghz来调节
           code ='1004b04b82d400890000'
       conditional_serial_communication(flag, code, port='COM5', baudrate=115200)
       cfg=cfg_xt
       flag=0
   for i in cfg:
       conditional_serial_communication(flag,i,port='COM5', baudrate=115200)
       time.sleep(0.1)
       uxr.set_uxr_state("RUN")
       time.sleep(0.1)
       uxr.auto_verticalscale(channal)
       time.sleep(0.1)
       uxr.clear_display()
       time.sleep(0.1)
       uxr.set_uxr_state("stop")
       freq = uxr.get_meas_val(channal, "FREQuency", "MEAN")
       freq = round(float(freq) * 1e-6, 3)
       temp_list = [chip_num, temp_volt, datarate, pattern,cfg_mode, code_mode,freq]
       # image_save_path="C:\\FK"     # 这个在路径是示波器里的路径
       uxr.save_screen_image(f"C:\\FK\\{chip_num}_{temp_volt}_{datarate}_{i}_{cfg_mode}_{code_mode}.png")  # path:C:\Users\Public\Documents\Infiniium
       time.sleep(2)
       print("频率是", freq)
       for i in range(len(temp_list)):
          ws.cell(row=last_row, column=i + 1, value=temp_list[i])
          wb.save("11.xlsx")
       last_row = last_row + 1
       temp_list.clear()
       flag=0
def conditional_serial_communication(flag,code,port='COM5', baudrate=115200):
    try:
        # 初始化串口
        ser = serial.Serial(
            port=port,
            baudrate=baudrate,
            bytesize=8,
            parity='N',
            stopbits=1,
            timeout=1
        )
        # 这个是为了让cfg top&phy的寄存器只走一次
        if flag==1:
            # cfg top通路
            top_sequences = [
            '1004b120482400000001',
            '1004b120485401000000',
            '1004b12048344A4B0000',
            '1004b1204848D3044110']
            for i, seq in enumerate(top_sequences, 1):
                byte_data = bytes.fromhex(seq)
                ser.write(byte_data)
                # 根据前缀决定是否读取响应
                if seq.startswith('1104'):
                    time.sleep(0.2)
                    response = ser.read_all()
                    if response:
                        print(f"接收数据: {response.hex()}")
                    else:
                        print("[警告] 未收到响应数据")
            # cfg phy 通路以及分频比
            phy_sequences = [
                '1004b04b80688c250000','1004b04b806884250000','1004b04b8124d6020000',
                '1004b04b811800200000', '1004b04b82d4B08E0000','1004b04b82cc01000000',
                '1004b04b82cc00000000','1004b04b81300f060000','1004b04bc03000000000']
            for i, seq in enumerate(phy_sequences, 1):
                byte_data = bytes.fromhex(seq)
                ser.write(byte_data)
                # 根据前缀决定是否读取响应
                if seq.startswith('1104'):
                    time.sleep(0.2)
                    response = ser.read_all()
                    if response:
                        print(f"接收数据: {response.hex()}")
                    else:
                        print("[警告] 未收到响应数据")

        hex_sequences=[code, '1004b04b82cc01000000','1004b04b82cc00000000']
        for i, seq in enumerate(hex_sequences, 1):
            byte_data = bytes.fromhex(seq)
            ser.write(byte_data)
            # 根据前缀决定是否读取响应
            if seq.startswith('1104'):
                time.sleep(0.2)
                response = ser.read_all()
                if response:
                      print(f"接收数据: {response.hex()}")
                else:
                    print("[警告] 未收到响应数据")


    except serial.SerialException as e:
        print(f"[错误] 串口通信异常: {e}")
    finally:
        if 'ser' in locals() and ser.is_open:
            ser.close()
            print("\n[状态] 串口已安全关闭")
def calculate_address(addr):
    """计算转换后的地址"""
    return 0xB0400000 + (0x1 << 19) + addr * 4
def initial_top(device):
    print("==============初始化top================")
    print("Verifying top...")
    init_top(device)

    '''
    [16:24:06.690] 读取成功 -> 地址: 0xb1204800, 数据: 0x20240821
    [16:24:06.691] 读取成功 -> 地址: 0xb1204804, 数据: 0x22334455
    [16:24:06.691] 读取成功 -> 地址: 0xb12048ac, 数据: 0x00000003
    [16:24:06.693] 读取成功 -> 地址: 0xb1204818, 数据: 0x00000001
    [16:24:06.693] 读取成功 -> 地址: 0xb1204848, 数据: 0x134d04d3
    [16:24:06.693] 读取成功 -> 地址: 0xb1204824, 数据: 0x02000000
    '''

    write_top_register(device, 0xb12048ac, 0x3)
    write_top_register(device, 0xb1204818, 0x1)
    write_top_register(device, 0xb1204848, 0x134d04d3)
    write_top_register(device, 0xb1204824, 0x2000000)

    print("0xb1204800:", read_top_register(device, 0xb1204800).hex())
    print("0xb1204804:", read_top_register(device, 0xb1204804).hex())
    print("0xb12048ac:", read_top_register(device, 0xb12048ac).hex())
    print("0xb1204818:", read_top_register(device, 0xb1204818).hex())
    print("0xb1204848:", read_top_register(device, 0xb1204848).hex())
    print("0xb1204824:", read_top_register(device, 0xb1204824).hex())

    print("==============top初始化成功================")
    time.sleep(0.1)
    print("==============*************************================")
def initial_serdes(device1):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0xf01e).hex()
    print("0xf01e:", value)
    if (value == "5983"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()
def configure_from_txt(txt_file, device):
    """
    从TXT文件读取并配置寄存器

    参数:
        txt_file: TXT文件路径
        device: 设备对象
        lane: 通道号
    """
    with open(txt_file, 'r') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()

        # 跳过空行和注释
        if not line or line.startswith('//') or line.startswith('#'):
            continue

        # 处理延迟
        if line.startswith('delay'):
            parts = line.split()
            if len(parts) > 1:
                delay_ms = int(parts[1])
                time.sleep(delay_ms / 1000.0)
            continue

        # 处理寄存器配置 (格式: 地址,值)
        if ',' in line:
            try:
                # 分割地址和值
                addr_str, value_str = line.split(',')

                # 转换为十六进制整数
                addr = int(addr_str.strip(), 16)
                value = int(value_str.strip(), 16)

                # 写入寄存器
                write_serdes_register(device, addr, value)
                time.sleep(0.005)
                value0=read_serdes_register(device, addr).hex()
                print(f"addr=0x{addr:04X},{value0}")
                print(f"addr=0x{addr:04X}, value=0x{value:04X}")


            except ValueError as e:
                print(f"解析行失败 '{line}': {e}")

if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    FTDI_CABLE_SERDES = b"A"  #
    FTDI_CABLE_TOP = b"B"  # fixme 设备名称，需要过滤出支持的设备列表供选择，或手动输入名称
    device = setup_device(FTDI_CABLE_TOP)
    device1 = setup_device(FTDI_CABLE_SERDES)
    initial_top(device)
    initial_serdes(device1)
    txt_file = '8G_all_inlp_prbs7.txt'   # ss的lowband使用8G_all_inlp_prbs7.txt
    configure_from_txt(txt_file, device1)
    close_serdes(device)
    close_serdes(device1)
    tk.messagebox.showinfo("JTAG配置完成，现在需要将J92(MFP0)接到示波器上，接好在点确认！！！")
    test()

