import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import re
import os
from datetime import datetime
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


class HexConverterApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.data = None
        self.file_path = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('十六进制转换工具 v1.0')
        self.setGeometry(200, 200, 1000, 700)

        # 设置应用图标
        self.setWindowIcon(QIcon(self.create_icon()))

        # 创建中心部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout(central_widget)

        # 创建菜单栏
        self.create_menu_bar()

        # 文件操作区域
        file_group = QGroupBox("文件操作")
        file_layout = QHBoxLayout()

        self.file_path_label = QLabel("未选择文件")
        self.file_path_label.setStyleSheet("border: 1px solid #ccc; padding: 5px;")
        self.file_path_label.setMinimumHeight(30)

        browse_btn = QPushButton("浏览文件")
        browse_btn.clicked.connect(self.browse_file)
        browse_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 8px;")

        file_layout.addWidget(self.file_path_label, 1)
        file_layout.addWidget(browse_btn)
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)

        # 参数设置区域
        params_group = QGroupBox("转换参数")
        params_layout = QGridLayout()

        # 编码格式选择
        params_layout.addWidget(QLabel("文本编码:"), 0, 0)
        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(['utf-8', 'gbk', 'ascii', 'latin-1'])
        params_layout.addWidget(self.encoding_combo, 0, 1)

        # 分隔符设置
        params_layout.addWidget(QLabel("分隔符:"), 0, 2)
        self.delimiter_combo = QComboBox()
        self.delimiter_combo.addItems(['空格', '逗号', '分号', '制表符', '换行符'])
        self.delimiter_combo.setCurrentIndex(0)
        params_layout.addWidget(self.delimiter_combo, 0, 3)

        # 数据起始行
        params_layout.addWidget(QLabel("数据起始行:"), 1, 0)
        self.start_row_spin = QSpinBox()
        self.start_row_spin.setRange(0, 1000)
        self.start_row_spin.setValue(0)
        params_layout.addWidget(self.start_row_spin, 1, 1)

        # 列数设置
        params_layout.addWidget(QLabel("列数:"), 1, 2)
        self.columns_spin = QSpinBox()
        self.columns_spin.setRange(1, 20)
        self.columns_spin.setValue(1)
        params_layout.addWidget(self.columns_spin, 1, 3)

        # 显示16进制原始数据
        self.show_hex_check = QCheckBox("在Excel中显示原始16进制数据")
        self.show_hex_check.setChecked(True)
        params_layout.addWidget(self.show_hex_check, 2, 0, 1, 4)

        params_group.setLayout(params_layout)
        main_layout.addWidget(params_group)

        # 预览区域
        preview_group = QGroupBox("数据预览")
        preview_layout = QVBoxLayout()

        # 创建标签页用于预览
        self.preview_tabs = QTabWidget()

        # 原始数据预览
        self.raw_text_edit = QTextEdit()
        self.raw_text_edit.setReadOnly(True)
        self.raw_text_edit.setMaximumHeight(150)
        self.preview_tabs.addTab(self.raw_text_edit, "原始数据")

        # 转换后数据预览
        self.converted_text_edit = QTextEdit()
        self.converted_text_edit.setReadOnly(True)
        self.converted_text_edit.setMaximumHeight(150)
        self.preview_tabs.addTab(self.converted_text_edit, "转换后数据")

        preview_layout.addWidget(self.preview_tabs)
        preview_group.setLayout(preview_layout)
        main_layout.addWidget(preview_group)

        # 图表区域
        chart_group = QGroupBox("图表预览")
        chart_layout = QVBoxLayout()

        # 创建matplotlib图形
        self.figure = Figure(figsize=(8, 4))
        self.canvas = FigureCanvas(self.figure)
        chart_layout.addWidget(self.canvas)

        chart_group.setLayout(chart_layout)
        main_layout.addWidget(chart_group)

        # 按钮区域
        button_layout = QHBoxLayout()

        convert_btn = QPushButton("转换数据")
        convert_btn.clicked.connect(self.convert_data)
        convert_btn.setStyleSheet("background-color: #2196F3; color: white; padding: 10px; font-weight: bold;")

        save_excel_btn = QPushButton("保存到Excel")
        save_excel_btn.clicked.connect(self.save_to_excel)
        save_excel_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; font-weight: bold;")

        save_chart_btn = QPushButton("保存图表")
        save_chart_btn.clicked.connect(self.save_chart)
        save_chart_btn.setStyleSheet("background-color: #FF9800; color: white; padding: 10px;")

        clear_btn = QPushButton("清除")
        clear_btn.clicked.connect(self.clear_all)
        clear_btn.setStyleSheet("background-color: #f44336; color: white; padding: 10px;")

        button_layout.addWidget(convert_btn)
        button_layout.addWidget(save_excel_btn)
        button_layout.addWidget(save_chart_btn)
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()

        main_layout.addLayout(button_layout)

        # 状态栏
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("就绪")

    def create_icon(self):
        """创建应用图标"""
        pixmap = QPixmap(32, 32)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)

        # 绘制图标
        painter.setPen(QPen(QColor(33, 150, 243), 2))
        painter.drawRect(4, 4, 24, 24)

        # 绘制Hex文字
        painter.setPen(QPen(QColor(76, 175, 80), 2))
        painter.setFont(QFont("Arial", 10, QFont.Bold))
        painter.drawText(QRect(4, 4, 24, 24), Qt.AlignCenter, "HEX")

        painter.end()
        return pixmap

    def create_menu_bar(self):
        """创建菜单栏"""
        menubar = self.menuBar()

        # 文件菜单
        file_menu = menubar.addMenu('文件')

        open_action = QAction('打开文件', self)
        open_action.setShortcut('Ctrl+O')
        open_action.triggered.connect(self.browse_file)
        file_menu.addAction(open_action)

        save_action = QAction('保存Excel', self)
        save_action.setShortcut('Ctrl+S')
        save_action.triggered.connect(self.save_to_excel)
        file_menu.addAction(save_action)

        file_menu.addSeparator()

        exit_action = QAction('退出', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # 工具菜单
        tools_menu = menubar.addMenu('工具')

        batch_action = QAction('批量转换', self)
        batch_action.triggered.connect(self.batch_convert)
        tools_menu.addAction(batch_action)

        # 帮助菜单
        help_menu = menubar.addMenu('帮助')

        about_action = QAction('关于', self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def browse_file(self):
        """浏览文件"""
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(
            self, "选择文本文件", "",
            "文本文件 (*.txt);;所有文件 (*)",
            options=options
        )

        if file_name:
            self.file_path = file_name
            self.file_path_label.setText(file_name)
            self.load_file_content()

    def load_file_content(self):
        """加载文件内容"""
        if not self.file_path:
            return

        try:
            encoding = self.encoding_combo.currentText()
            with open(self.file_path, 'r', encoding=encoding, errors='ignore') as file:
                content = file.read()

            # 显示原始数据预览
            self.raw_text_edit.setPlainText(content[:1000] + ("..." if len(content) > 1000 else ""))
            self.statusBar.showMessage(f"已加载文件: {os.path.basename(self.file_path)}", 3000)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取文件失败: {str(e)}")

    def convert_data(self):
        """转换数据"""
        if not self.file_path:
            QMessageBox.warning(self, "警告", "请先选择文件")
            return

        try:
            # 获取参数
            encoding = self.encoding_combo.currentText()
            start_row = self.start_row_spin.value()
            num_columns = self.columns_spin.value()

            # 获取分隔符
            delimiter_map = {
                '空格': r'\s+',
                '逗号': ',',
                '分号': ';',
                '制表符': '\t',
                '换行符': '\n'
            }
            delimiter = delimiter_map[self.delimiter_combo.currentText()]

            # 读取文件
            with open(self.file_path, 'r', encoding=encoding, errors='ignore') as file:
                lines = file.readlines()

            # 跳过指定行数
            lines = lines[start_row:]

            # 解析数据
            hex_data = []
            decimal_data = []

            for line_num, line in enumerate(lines, start=start_row + 1):
                line = line.strip()
                if not line:
                    continue

                # 分割数据
                if delimiter == r'\s+':
                    parts = re.split(r'\s+', line)
                else:
                    parts = line.split(delimiter)

                # 只取指定列数的数据
                parts = parts[:num_columns]

                hex_row = []
                decimal_row = []

                for part in parts:
                    part = part.strip()
                    if not part:
                        hex_row.append('')
                        decimal_row.append(np.nan)
                        continue

                    # 尝试转换为16进制（去除0x前缀）
                    clean_part = part.strip()
                    if clean_part.lower().startswith('0x'):
                        clean_part = clean_part[2:]

                    try:
                        # 尝试解析为16进制整数
                        decimal_val = int(clean_part, 16)
                        hex_row.append(f"0x{clean_part.upper()}")
                        decimal_row.append(decimal_val)
                    except ValueError:
                        # 如果不是有效的16进制，尝试作为普通数字
                        try:
                            decimal_val = float(clean_part)
                            hex_row.append(clean_part)
                            decimal_row.append(decimal_val)
                        except:
                            hex_row.append(clean_part)
                            decimal_row.append(np.nan)

                hex_data.append(hex_row)
                decimal_data.append(decimal_row)

            # 转换为DataFrame
            hex_columns = [f"Hex_Col{i + 1}" for i in range(num_columns)]
            decimal_columns = [f"Dec_Col{i + 1}" for i in range(num_columns)]

            self.hex_df = pd.DataFrame(hex_data, columns=hex_columns)
            self.decimal_df = pd.DataFrame(decimal_data, columns=decimal_columns)

            # 创建包含原始16进制和转换后十进制的完整DataFrame
            combined_data = []
            for i in range(len(hex_data)):
                row = []
                for j in range(num_columns):
                    row.append(hex_data[i][j])
                    row.append(decimal_data[i][j])
                combined_data.append(row)

            # 创建列名
            combined_columns = []
            for i in range(num_columns):
                combined_columns.append(f"原始数据_列{i + 1}")
                combined_columns.append(f"十进制_列{i + 1}")

            self.combined_df = pd.DataFrame(combined_data, columns=combined_columns)

            # 显示转换后数据预览
            preview_text = f"共转换 {len(decimal_data)} 行数据，{num_columns} 列\n\n"
            preview_text += "前10行数据预览:\n"
            preview_text += self.combined_df.head(10).to_string()
            self.converted_text_edit.setPlainText(preview_text)

            # 绘制图表
            self.plot_data(decimal_data, num_columns)

            self.data = decimal_data
            self.statusBar.showMessage(f"转换完成: {len(decimal_data)} 行数据已处理", 5000)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"数据转换失败: {str(e)}")
            import traceback
            traceback.print_exc()

    def plot_data(self, data, num_columns):
        """绘制折线图"""
        self.figure.clear()

        if not data:
            self.canvas.draw()
            return

        ax = self.figure.add_subplot(111)

        # 将数据转换为numpy数组
        data_array = np.array(data)

        # 绘制每一列的数据
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD', '#98D8C8']

        for i in range(min(num_columns, data_array.shape[1])):
            column_data = data_array[:, i]
            # 过滤NaN值
            valid_indices = ~np.isnan(column_data)
            if np.any(valid_indices):
                x = np.where(valid_indices)[0]
                y = column_data[valid_indices]
                ax.plot(x, y, marker='o', markersize=3, linewidth=1.5,
                        color=colors[i % len(colors)], label=f'列 {i + 1}')

        ax.set_xlabel('数据点序号')
        ax.set_ylabel('数值')
        ax.set_title('十六进制转换数据折线图')
        ax.legend(loc='best')
        ax.grid(True, alpha=0.3)

        # 自动调整布局
        self.figure.tight_layout()
        self.canvas.draw()

    def save_to_excel(self):
        """保存到Excel文件"""
        if not hasattr(self, 'combined_df') or self.combined_df.empty:
            QMessageBox.warning(self, "警告", "没有数据可保存，请先转换数据")
            return

        try:
            options = QFileDialog.Options()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"hex_converted_{timestamp}.xlsx"

            file_name, _ = QFileDialog.getSaveFileName(
                self, "保存Excel文件", default_name,
                "Excel文件 (*.xlsx);;所有文件 (*)",
                options=options
            )

            if file_name:
                if not file_name.endswith('.xlsx'):
                    file_name += '.xlsx'

                # 创建Excel写入器
                with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                    # 保存转换后的数据
                    self.combined_df.to_excel(writer, sheet_name='转换数据', index=False)

                    # 添加汇总统计信息
                    summary_data = []
                    num_columns = len([col for col in self.combined_df.columns if '十进制' in col])

                    for i in range(num_columns):
                        decimal_col = f"十进制_列{i + 1}"
                        if decimal_col in self.combined_df.columns:
                            col_data = self.combined_df[decimal_col].dropna()
                            if not col_data.empty:
                                summary_data.append([
                                    f"列{i + 1}",
                                    len(col_data),
                                    col_data.min(),
                                    col_data.max(),
                                    col_data.mean(),
                                    col_data.std()
                                ])

                    summary_df = pd.DataFrame(
                        summary_data,
                        columns=['列名', '数据点数', '最小值', '最大值', '平均值', '标准差']
                    )
                    summary_df.to_excel(writer, sheet_name='统计信息', index=False)

                    # 添加元数据
                    meta_data = {
                        '项目': ['源文件', '转换时间', '总行数', '总列数', '编码格式', '分隔符'],
                        '值': [
                            self.file_path or '未知',
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            len(self.combined_df),
                            num_columns,
                            self.encoding_combo.currentText(),
                            self.delimiter_combo.currentText()
                        ]
                    }
                    meta_df = pd.DataFrame(meta_data)
                    meta_df.to_excel(writer, sheet_name='文件信息', index=False)

                # 将图表保存到Excel
                self.add_chart_to_excel(file_name)

                self.statusBar.showMessage(f"数据已保存到: {file_name}", 5000)
                QMessageBox.information(self, "成功", f"数据已成功保存到:\n{file_name}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存Excel文件失败: {str(e)}")

    def add_chart_to_excel(self, excel_file):
        """将图表添加到Excel文件"""
        try:
            # 保存图表为临时图片
            temp_chart_path = "temp_chart.png"
            self.figure.savefig(temp_chart_path, dpi=150, bbox_inches='tight')

            # 将图片插入Excel
            wb = openpyxl.load_workbook(excel_file)

            # 创建图表工作表
            if '图表' in wb.sheetnames:
                ws_chart = wb['图表']
            else:
                ws_chart = wb.create_sheet('图表')

            # 清除现有内容
            ws_chart.delete_rows(1, ws_chart.max_row)
            ws_chart.delete_cols(1, ws_chart.max_column)

            # 添加图表标题
            ws_chart['A1'] = '数据可视化图表'
            ws_chart['A1'].font = openpyxl.styles.Font(size=14, bold=True)

            # 插入图片
            from openpyxl.drawing.image import Image
            img = Image(temp_chart_path)
            img.anchor = 'A3'  # 从A3单元格开始
            ws_chart.add_image(img)

            # 保存并关闭
            wb.save(excel_file)
            wb.close()

            # 删除临时图片
            if os.path.exists(temp_chart_path):
                os.remove(temp_chart_path)

        except Exception as e:
            print(f"将图表添加到Excel时出错: {e}")

    def save_chart(self):
        """单独保存图表"""
        if not hasattr(self, 'combined_df') or self.combined_df.empty:
            QMessageBox.warning(self, "警告", "没有图表可保存，请先转换数据")
            return

        try:
            options = QFileDialog.Options()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"chart_{timestamp}.png"

            file_name, _ = QFileDialog.getSaveFileName(
                self, "保存图表", default_name,
                "PNG图像 (*.png);;JPEG图像 (*.jpg);;PDF文件 (*.pdf)",
                options=options
            )

            if file_name:
                # 根据文件扩展名确定格式
                if file_name.endswith('.pdf'):
                    self.figure.savefig(file_name, format='pdf', bbox_inches='tight')
                elif file_name.endswith('.jpg') or file_name.endswith('.jpeg'):
                    self.figure.savefig(file_name, format='jpeg', dpi=300, bbox_inches='tight')
                else:
                    self.figure.savefig(file_name, format='png', dpi=300, bbox_inches='tight')

                self.statusBar.showMessage(f"图表已保存到: {file_name}", 5000)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存图表失败: {str(e)}")

    def batch_convert(self):
        """批量转换"""
        QMessageBox.information(self, "信息", "批量转换功能正在开发中")

    def clear_all(self):
        """清除所有"""
        reply = QMessageBox.question(
            self, '确认',
            '确定要清除所有数据吗？',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            self.file_path = None
            self.file_path_label.setText("未选择文件")
            self.raw_text_edit.clear()
            self.converted_text_edit.clear()
            self.figure.clear()
            self.canvas.draw()
            self.statusBar.showMessage("已清除所有数据", 3000)

    def show_about(self):
        """显示关于对话框"""
        about_text = """
        <h2>十六进制转换工具 v1.0</h2>
        <p>功能说明:</p>
        <ul>
          <li>读取包含16进制数据的文本文件</li>
          <li>将16进制值转换为10进制</li>
          <li>保存转换结果到Excel文件</li>
          <li>绘制数据折线图</li>
          <li>支持多种分隔符和编码格式</li>
        </ul>
        <p>使用方法:</p>
        <ol>
          <li>点击"浏览文件"选择文本文件</li>
          <li>设置转换参数（编码、分隔符等）</li>
          <li>点击"转换数据"进行转换</li>
          <li>点击"保存到Excel"保存结果</li>
          <li>点击"保存图表"单独保存图表</li>
        </ol>
        <p>© 2023 版权所有</p>
        """
        QMessageBox.about(self, "关于", about_text)


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    # 设置应用程序样式
    app.setStyleSheet("""
        QMainWindow {
            background-color: #f5f5f5;
        }
        QGroupBox {
            font-weight: bold;
            border: 2px solid #cccccc;
            border-radius: 5px;
            margin-top: 10px;
            padding-top: 10px;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px 0 5px;
        }
    """)

    window = HexConverterApp()
    window.show()

    sys.exit(app.exec_())


if __name__ == '__main__':
    main()