import winsound
from ftd2jtag.jtag_reg_op import  *
from ftd2jtag.ftd2jtag import *
import openpyxl
from instrument import DM3058E
import datetime

test_programs_config_p = {
    "rx_vref": {
        "registers": [(0xe0d4, 0x0080), (0xe0cc, 0x0008)],
    },
    "vreg_vco": {
        "registers": [(0xe0d4, 0x0080), (0xe0cb, 0x0004)],
    },
    "vreg_clk": {
        "registers": [(0xe0d4, 0x0080), (0xe0cb, 0x0020)],
    },
    "vreg_vro": {
        "registers": [(0xe0d4, 0x0080), (0xe0c7, 0x0008)],
    },
    "vreg_dfe": {
        "registers": [(0xe0d4, 0x0080), (0xe0c7, 0x0020)],
    },
    "vreg_dfe_by": {
        "registers": [(0xe0d4, 0x0080), (0xe0c7, 0x0010)],
    },
    "vdcc_i_p": {
        "registers": [(0xe0d4, 0x0080), (0xe0d3, 0x0004)],
    },
    "vdcc_q_p": {
        "registers": [(0xe0d4, 0x0080), (0xe0d3, 0x0008)],
    }
,
    "samp_clk_i_p": {
        "registers": [(0xe0d4, 0x0080), (0xe0d1,0x0001)],
    },
    "samp_clk_q_p": {
        "registers": [(0xe0d4, 0x0080), (0xe0d1,0x0002)],
    },
    "vp_p": {
        "registers": [(0xe0d4, 0x0080), (0xe0d3, 0x0008)],
    },
    "slowreg_vref_p": {
        "registers": [(0xe0d4, 0x0000), (0xf054, 0x0001)],
    },
    "fastreg_vref": {
        "registers": [(0xe0d4,0x0000), (0xf054, 0x0004)],
    },
    "vbg": {
        "registers": [(0xe0d4,0x0000), (0xf054,0x0080)],
    },
    "vp_h_p": {
        "registers": [(0xe0d4,0x0000), (0xf054,0x0020)],
    }

}

test_programs_config_n = {
    "rx car vco vosc": {
        "registers": [(0xe0d4, 0x0080), (0xe0cb,0x0001)],
    },
    "vco_gd_n": {
        "registers": [(0xe0d4, 0x0080), (0xe0cb,0x0002)],
    },
    "vdcc i_m": {
        "registers": [(0xe0d4, 0x0080), (0xe0d3,0x0004)],
    },
    "vdcc_g_m": {
        "registers": [(0xe0d4, 0x0080), (0xe0d3,0x0008)],
    },
    "samp_clk i_m": {
        "registers": [(0xe0d4, 0x0080), (0xe0d1,0x0001)],
    },
    "samp_clk_g_m": {
        "registers": [(0xe0d4, 0x0080), (0xe0d1,0x0002)],
    }

}

def test():
   wb=openpyxl.load_workbook("DC.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   chip_num = "TT01"
   temp_volt = "NTNV"
   # 检查硬件先测P端还是N端
   polarity = "P"  # "N"
   print(f"目前测试: {polarity}")
   value_list = []
   DM3058E.init()
   DM3058E.reset()
   time.sleep(1)
   for i in range(4):
        clear_script(device1,i)
   if(polarity=="P"):
       temp_list = [chip_num, temp_volt, polarity]
       for test_program, config in test_programs_config_p.items():
           for i in range(4):
               # test lane
               value_list.append(i)
               value_list.append(test_program)
               print("目前测试lane",i,test_program)
               for reg_addr, reg_value in config["registers"]:
                    write_serdes_lane_register(device1, i, reg_addr, reg_value)
               volt=DM3058E.get_voltage()
               time.sleep(0.1)
               print("电压是",volt,"v")
               value_list.append(volt)
               time.sleep(0.1)
               clear_script(device1, i)
               for i in range(len(temp_list)):
                    ws.cell(row=last_row, column=i + 1, value=temp_list[i])
               for i in range(len(value_list)):
                    ws.cell(row=last_row, column=i + 4, value=value_list[i])
               wb.save("P5_DC.xlsx")
               save_test_data_to_txt("DC", temp_list, value_list)
               last_row=last_row+1
               value_list.clear()
   else:
        temp_list = [chip_num, temp_volt, polarity]
        for test_program, config in test_programs_config_n.items():
            for i in range(4):
                # test lane
                value_list.append(i)
                value_list.append(test_program)
                print("目前测试lane", i, test_program)
                for reg_addr, reg_value in config["registers"]:
                    write_serdes_lane_register(device1, i, reg_addr, reg_value)
                volt = DM3058E.get_voltage()
                print("电压是", volt, "v")
                value_list.append(volt)
                clear_script(device1, i)
                time.sleep(0.1)
                for i in range(len(temp_list)):
                    ws.cell(row=last_row, column=i + 1, value=temp_list[i])
                for i in range(len(value_list)):
                    ws.cell(row=last_row, column=i + 4, value=value_list[i])
                wb.save("P5_DC.xlsx")
                save_test_data_to_txt("DC", temp_list, value_list)
                last_row = last_row + 1
                value_list.clear()

def initial_top(device):
    print("==============初始化top================")
    print("Verifying top...")
    init_top(device)

    '''
    [16:24:06.690] 读取成功 -> 地址: 0xb1204800, 数据: 0x20240821
    [16:24:06.691] 读取成功 -> 地址: 0xb1204804, 数据: 0x22334455
    [16:24:06.691] 读取成功 -> 地址: 0xb12048ac, 数据: 0x00000003
    [16:24:06.693] 读取成功 -> 地址: 0xb1204818, 数据: 0x00000001
    [16:24:06.693] 读取成功 -> 地址: 0xb1204848, 数据: 0x134d04d3
    [16:24:06.693] 读取成功 -> 地址: 0xb1204824, 数据: 0x02000000
    '''

    write_top_register(device, 0xb12048ac, 0x3)
    write_top_register(device, 0xb1204818, 0x1)
    write_top_register(device, 0xb1204848, 0x134d04d3)
    write_top_register(device, 0xb1204824, 0x2000000)

    print("0xb1204800:", read_top_register(device, 0xb1204800).hex())
    print("0xb1204804:", read_top_register(device, 0xb1204804).hex())
    print("0xb12048ac:", read_top_register(device, 0xb12048ac).hex())
    print("0xb1204818:", read_top_register(device, 0xb1204818).hex())
    print("0xb1204848:", read_top_register(device, 0xb1204848).hex())
    print("0xb1204824:", read_top_register(device, 0xb1204824).hex())

    print("==============top初始化成功================")
    time.sleep(0.1)
    print("==============*************************================")

def initial_serdes(device1):
    print("Verifying serdes...")
    init_serdes(device1)
    value = read_serdes_register(device1, 0xf01e).hex()
    print("0xf01e:", value)
    if (value == "5983"):
        print("环境没有问题！！！")
    else:
        print("环境有问题！！！")
        exit()

def clear_script(device1,lane):
    write_serdes_lane_register(device1, lane, 0xe0cc, 0x0001)
    write_serdes_lane_register(device1, lane, 0xe0cb, 0x0000)
    write_serdes_lane_register(device1, lane, 0xe0c7, 0x0000)
    write_serdes_lane_register(device1, lane, 0xe0d3, 0x0000)
    write_serdes_lane_register(device1, lane, 0xe0d1, 0x0000)
    write_serdes_lane_register(device1, lane, 0xe0d4, 0x0000)
    write_serdes_register(device1, 0xf054, 0x0000)

test_cfg1={
    "cfg1": {
        "registers": [(0xe01f,0x0005),(0xe01b,0x0008),(0xc003,0x022d),(0xc000,0x100a),
                        (0xc004,0x1017),(0xe004,0x0101),(0xc001,0x0c00),(0xc006,0x0010),
                        (0xc017,0x01c0),(0xe01d,0xd1d0),(0xe01c,0x2040)],
    }
}
test_cfg2={
    "cfg2": {
        "registers": [(0xc0de,0x0012),(0xc007,0x48e2),(0xc018,0x2550),(0xc01f,0x1f10),
                        (0xe01e,0xb019),(0xe01a,0x2408),(0xe01e,0xb009),(0xe01a,0x2400),
                        (0xe034,0x0008),(0xc064,0x0007),(0xe01a,0x2400),(0xe01a,0x2404)
                      ,(0xe04e,0x0008),(0xe04e,0x0018),(0xe04e,0x0018),(0xe04e,0x0018)
                      ,(0xe04e,0x0018),(0xe034,0x0008),(0xc064,0x0007),(0xe01a,0x2400)
                      ,(0xe01a,0x2404),(0xe04e,0x0008),(0xe04e,0x0018),(0xe04e,0x0018)
                      ,(0xe04e,0x0018),(0xe04e,0x0018),(0xe034,0x0018),(0xe034,0x0018)
                      ,(0xe034,0x0018)],
    }
}
def inlp():
    for i in range(4):
        write_serdes_lane_register(device1,i, 0xc009,0x000a)
        time.sleep(0.1)
    for cfg1,config in test_cfg1.items():
        for i in range(4):
            for reg_addr, reg_value in config["registers"]:
                write_serdes_lane_register(device1, i, reg_addr, reg_value)
    write_serdes_register(device1, 0xf01c,0x1000)
    write_serdes_register(device1, 0xd02b,0x000c)
    write_serdes_register(device1, 0xf032,0x0008)
    write_serdes_register(device1, 0xf029,0x0045)
    write_serdes_register(device1, 0xf012,0xa3a6)
    write_serdes_register(device1, 0xf016,0x4506)
    write_serdes_register(device1, 0xf016,0x4507)
    write_serdes_register(device1, 0xf01b,0x8000)
    for cfg2,config in test_cfg2.items():
        for i in range(4):
            for reg_addr, reg_value in config["registers"]:
                write_serdes_lane_register(device1, i, reg_addr, reg_value)

def save_test_data_to_txt(filename, temp_list, value_list, timestamp=None):
    """
    保存测试数据到TXT文件

    参数:
        filename: 文件名
        temp_list: [芯片编号, 温度/电压, 极性]
        value_list: [测试通道, 测试程序, 电压值]
        timestamp: 时间戳（可选）
    """
    import datetime

    # 如果没有提供时间戳，生成当前时间
    if timestamp is None:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 合并所有数据
    all_data = temp_list + value_list + [timestamp]

    # 转换为字符串
    data_line = ','.join(str(item) for item in all_data)

    # 检查文件是否存在，决定是否添加表头
    if not os.path.exists(filename):
        # 文件不存在，创建并添加表头
        header = "芯片编号,温度/电压,极性,测试通道,测试程序,电压值,时间戳"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(header + '\n')
            f.write(data_line + '\n')
    else:
        # 文件存在，追加数据
        with open(filename, 'a', encoding='utf-8') as f:
            f.write(data_line + '\n')

    print(f"数据已保存到: {filename}")
    return True

def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")



if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    FTDI_CABLE_SERDES = b"A"  #
    FTDI_CABLE_TOP = b"B"  # fixme 设备名称，需要过滤出支持的设备列表供选择，或手动输入名称
    device = setup_device(FTDI_CABLE_TOP)
    device1 = setup_device(FTDI_CABLE_SERDES)
    initial_top(device)
    initial_serdes(device1)
    inlp()
    test()
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    print(f"总耗时: {duration}")
    play_sound_test_complete()