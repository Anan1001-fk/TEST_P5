import time
from idlelib.autocomplete import FORCE

import openpyxl
import uxr0134a

def test():
   wb=openpyxl.load_workbook("P4/11.xlsx")
   ws=wb.worksheets[1]
   last_row=ws.max_row+1
   address = "TCPIP0::169.254.202.222::inst0::INSTR"
   uxr=uxr0134a.Uxr0134a(address)
   channal=2
   # vpp = uxr.get_meas_val(channal, "VPP")
   # vpp = vpp * 1000
   uxr.set_uxr_state("stop")
   time.sleep(0.1)
   risetime = uxr.get_meas_val(channal, "RISetime", "MEAN")
   risetime = round(float(risetime) * 1e12, 3)
   falltime = uxr.get_meas_val(channal, "FALLtime", "MEAN")
   falltime = round(float(falltime) * 1e12, 3)
   eyehight = uxr.get_eh_value_mv()
   eyewidth = uxr.get_ew_value_ps()
   jitter = uxr.get_tj()
   tj = jitter[0]
   tj = round(float(tj) * 1e12, 3)
   rj = jitter[1]
   rj = round(float(rj) * 1e12, 3)
   isi = jitter[7]
   isi = round(float(isi) * 1e12, 3)
   dcd = jitter[6]
   dcd = round(float(dcd) * 1e12, 3)
   temp_list = ["SS06", "LTLV", "16Gbps", "LANE1_PRBS31_7%", risetime, falltime, eyehight, eyewidth, tj, rj, isi,
                dcd]
   uxr.save_screen_image("SS06_16Gbps_PRBS31_LTLV_LANE1_7%.png")  # path:C:\Users\Public\Documents\Infiniium
   for i in range(len(temp_list)):
    ws.cell(row=last_row, column=i + 1, value=temp_list[i])
    wb.save("11.xlsx")
if __name__=="__main__":
    test()
    # address = "TCPIP0::169.254.202.222::inst0::INSTR"
    # uxr = uxr0134a.Uxr0134a(address)
    # uxr.save_screen_image("FF06_16Gbps_HTHV_LaneSkew.png")