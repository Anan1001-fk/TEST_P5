import time
import dp832
import openpyxl
import uxr0134a
import serial.tools
def test():
   wb=openpyxl.load_workbook("P4/11.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   address = "USB0::0x1AB1::0x0E11::DP8C162851463::INSTR"
   uxr0134a.Uxr0134a.__init__(res_name=address)
   for i in range(1,256):
       write_reg(0xf0,i)
       v=dp832.get_voltage(1)

   # dp832.init(address=address)
   # dp832.reset()
   # time.sleep(0.01)
   # dp832.set_voltage_current(channel=1, voltage=5, current=1)
   # while True:
   #  dp832.output(1,"ON")
   #  time.sleep(1)
   #  volt=dp832.get_voltage(1)
   #  print(volt)
   #  current = dp832.get_current(1)
   #  print(current)
   temp_list = ["TT22", "NT NV", volt, current]
   for i in range(len(temp_list)):
    ws.cell(row=last_row, column=i + 1, value=temp_list[i])
    wb.save("11.xlsx")
    time.sleep(1)
if __name__=="__main__":
    test()
