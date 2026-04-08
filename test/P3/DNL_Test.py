import os
import time
from instrument import FSWP
import winsound
import datetime
import openpyxl
from retimer_sdk_drive import BIF_NAMES, LINK_NAMES, LTSSM_NAMES, RetimerSDKDrive


def test():
   wb=openpyxl.load_workbook("ki_test.xlsx")
   ws=wb.worksheets[0]
   last_row=ws.max_row+1
   chip_num = "TT01"
   temp_volt = "NTNV"
   temp_list=[chip_num,temp_volt]
   value_list=[]
   for i in range(0,7,2):
       print("目前测试ki",i)
       value_list.append(i)
       initial(ki_code=i)
       time.sleep(0.1)
       FSWP.run()
       FSWP.stop()
       FSWP.peaking_search()
       x = FSWP.get_mark_X()
       y = FSWP.get_mark_y()
       print(x)
       print(y)
       value_list.append(x)
       value_list.append(y)
       for i in range(len(temp_list)):
           ws.cell(row=last_row, column=i + 1, value=temp_list[i])
       for i in range(len(value_list)):
           ws.cell(row=last_row, column=i + 3, value=value_list[i])
       wb.save("ki_test.xlsx")
       save_test_data_to_txt("ki_test", temp_list, value_list)
       last_row = last_row + 1
       value_list.clear()
def initial(ki_code):
    sdk.write_reg32(handle,0x58004000,0x0004)
    sdk.write_reg32(handle, 0x5800C000, 0x1F10)
    sdk.write_reg32(handle, 0x5800C004, 0x1FE0)
    time.sleep(0.01)
    sdk.write_reg32(handle, 0x5800C01C, 0xA550)
    sdk.write_reg32(handle, 0x5800C020, 0x9629)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x5800C458, 0x0AAA)
    sdk.write_reg32(handle, 0x5800C060, 0x0784)
    sdk.write_reg32(handle, 0x5800C064, 0x0f04)
    sdk.write_reg32(handle, 0x58008004, 0x085C)
    sdk.write_reg32(handle, 0x58008008, 0x0000)
    sdk.write_reg32(handle, 0x5800800C, 0x0002)
    sdk.write_reg32(handle, 0x58008010, 0x0867)
    sdk.write_reg32(handle, 0x58008014, 0x0000)
    sdk.write_reg32(handle, 0x58008018, 0x0002)
    sdk.write_reg32(handle, 0x58008020, 0x002A)
    sdk.write_reg32(handle, 0x5800C304, 0x0014)
    sdk.write_reg32(handle, 0x5800C304, 0x0015)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x5800C058, 0x0FEA)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x5800C058, 0x0FAA)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x5800C058, 0x0FBA)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x5800C058, 0x0FAA)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x58000010, 0x2329)
    sdk.write_reg32(handle, 0x5800001C, 0x1315)
    sdk.write_reg32(handle, 0x58000024, 0x45)
    sdk.write_reg32(handle, 0x58004014, 0x7400)
    sdk.write_reg32(handle, 0x58004014, 0x7401)
    sdk.write_reg32(handle, 0x58004124, 0x107)
    time.sleep(0.01)
    sdk.write_reg32(handle, 0x58004014, 0x7400)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x58004124, 0x103)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x58004124, 0x3)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x58004014, 0x7404)
    time.sleep(0.02)
    sdk.write_reg32(handle, 0x58004014, 0x7400)
    ki_code=ki_code*16
    sdk.write_reg32(handle, 0x58004164, ki_code)   # default=0
    print("当前的ki_code是", ki_code)
    sdk.write_reg32(handle, 0x58004398, 0x01F3)
    sdk.write_reg32(handle, 0x58004394, 1)
    time.sleep(0.01)
    sdk.write_reg32(handle, 0x50030308, 0x33)
    sdk.write_reg32(handle, 0x50030138, 0x180)
    sdk.write_reg32(handle, 0x50030100, 0x15596)
    sdk.write_reg32(handle, 0x50030124, 0x104C1)
    sdk.write_reg32(handle, 0x50030140, 0x535555)
    sdk.write_reg32(handle, 0x5800414C, 0x60F)








def save_test_data_to_txt(filename, temp_list, timestamp=None):
    """
    保存测试数据到TXT文件

    参数:
        filename: 文件名
        temp_list: [芯片编号, 温度/电压, 极性]
        value_list: [测试通道, 测试程序, 电压值]
        timestamp: 时间戳（可选）
    """
    import datetime

    # 如果没有提供时间戳，生成当前时间
    if timestamp is None:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 合并所有数据
    all_data = temp_list  + [timestamp]

    # 转换为字符串
    data_line = ','.join(str(item) for item in all_data)

    # 检查文件是否存在，决定是否添加表头
    if not os.path.exists(filename):
        # 文件不存在，创建并添加表头
        header = "芯片编号,温度/电压,极性,测试通道,测试程序,电压值,时间戳"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(header + '\n')
            f.write(data_line + '\n')
    else:
        # 文件存在，追加数据
        with open(filename, 'a', encoding='utf-8') as f:
            f.write(data_line + '\n')

    print(f"数据已保存到: {filename}")
    return True

def play_sound_test_complete():
    """播放测试完成的声音提示（Windows）"""
    try:
        # 播放系统提示音
        winsound.MessageBeep()

        # 或者播放自定义频率和时长的声音
        frequency = 1000  # 频率（赫兹）
        duration = 1000  # 时长（毫秒）
        winsound.Beep(frequency, duration)

        print("测试完成！")
    except Exception as e:
        print(f"播放声音失败: {e}")



if __name__=="__main__":
    test_start_time = datetime.datetime.now()
    print(f"测试开始时间: {test_start_time}")
    """主测试函数"""
    print("=" * 60)
    print("Retimer SDK Drive 测试程序")
    print("=" * 60)
    # 创建SDK Drive实例
    sdk = RetimerSDKDrive()
    # 设置日志级别 (2=INFO)
    sdk.lib.retimer_set_log_level(2)
    # 1. 枚举设备
    print("\n1. 枚举设备...")
    devices = sdk.enumerate_devices()
    print(f"   找到 {len(devices)} 个设备:")
    for i, dev in enumerate(devices):
        print(f"   [{i}] Type={dev['type']}, Port={dev['port']}, "
              f"Status={dev['status']}, Desc={dev['description']}")
    if len(devices) == 0:
        print("   未找到设备，退出测试")
    # 2. 打开设备
    print("\n2. 打开设备...")
    handle = sdk.open_device(
        adapter_type=devices[0]['type'],
        port=devices[0]['port'],
        slave_addr=0x20,
        bitrate_khz=800,
        pec_enable=True,
        chip_version=1  # 1=NTO
    )
    print(f"   设备已打开: handle={handle}")
    test()
    test_end_time = datetime.datetime.now()
    print(f"测试结束时间: {test_end_time}")
    duration = test_end_time - test_start_time
    print(f"总耗时: {duration}")
    play_sound_test_complete()